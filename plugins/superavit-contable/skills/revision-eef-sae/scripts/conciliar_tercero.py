#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
conciliar_tercero.py — Método categórico de detección de diferencias para cuentas
con control por tercero (CxC, CxP, anticipos de clientes/proveedores, CxC empleados).
Muele el mayor de la cuenta + el reporte de módulo (Hoja 2 por RUC) y, opcionalmente,
el reporte de módulo al cierre del periodo anterior (apertura por tercero). Devuelve un
RESUMEN corto: el descuadre global partido en apertura + periodo, y los terceros
descuadrados. SOLO LEE. Implementa los Pasos 0-3 de «Detección de diferencias» del cerebro.

Uso:
  python conciliar_tercero.py --mayor MAYOR.csv --cuenta 20110 --modulo MOD.xlsx \
        [--apertura APE_CIERRE_ANTERIOR.xlsx] [--anio 2026] [--detalle OUT.xlsx]
"""
import argparse, csv, re, unicodedata
from datetime import date, datetime
from pathlib import Path

def num(s):
    if s is None: return 0.0
    s=str(s).strip()
    if s in ("","-"): return 0.0
    neg=s.startswith("-") or s.startswith("(")
    s=s.lstrip("-(").rstrip(")").replace(".","").replace(",",".")
    s=re.sub(r"[^0-9.\-]","",s)
    try: v=float(s)
    except ValueError: return 0.0
    return -v if neg else v

def norm(t):
    t=unicodedata.normalize("NFKD",str(t)).encode("ascii","ignore").decode().upper()
    return re.sub(r"\s+"," ",t).strip()

def canon(t):
    toks=norm(t).split()
    out=[]
    for w in toks:
        if w not in out: out.append(w)
    return " ".join(out[:4])

def pdate(v):
    if isinstance(v,(datetime,date)): return v if isinstance(v,date) else v.date()
    for fmt in ("%d/%m/%Y","%Y-%m-%d","%d-%m-%Y"):
        try: return datetime.strptime(str(v).strip(),fmt).date()
        except: pass
    return None

def cargar_mayor_cuenta(path, cuenta):
    raw=Path(path).read_bytes()
    for enc in ("utf-8-sig","utf-8","latin-1"):
        try: text=raw.decode(enc); break
        except UnicodeDecodeError: continue
    rows=list(csv.reader(text.splitlines(), delimiter=";"))
    H=[norm(h) for h in rows[0]]
    def col(*names):
        for n in names:
            for i,h in enumerate(H):
                if norm(n)==h or norm(n) in h: return i
        return None
    ic={"cod":col("CODIGO CUENTA","CODIGO"),"ben":col("BENEFICIARIO"),
        "fecha":col("FECHA"),"debe":col("DEBE"),"haber":col("HABER")}
    apertura=0.0; periodo={}; mov_periodo=0.0
    for r in rows[1:]:
        if not r or len(r)<=ic["cod"]: continue
        if str(r[ic["cod"]]).strip()!=str(cuenta): continue
        d=num(r[ic["debe"]]); h=num(r[ic["haber"]]); s=d+h
        ben=str(r[ic["ben"]]).strip() if ic["ben"] is not None else ""
        if not ben:
            apertura+=s
        else:
            periodo[canon(ben)]=periodo.get(canon(ben),0.0)+s
            mov_periodo+=s
    return round(apertura,2), {k:round(v,2) for k,v in periodo.items()}, round(mov_periodo,2)

def cargar_modulo(path):
    import openpyxl
    wb=openpyxl.load_workbook(path,data_only=True)
    ws=None; hdr=None
    for sh in wb.worksheets:
        for i,row in enumerate(sh.iter_rows(values_only=True),1):
            vals=[norm(v) for v in row if v is not None]
            if any("RUC" in v or "CEDULA" in v for v in vals) and any("PENDIENTE" in v for v in vals):
                ws=sh; hdr=i; HR=[norm(v) if v is not None else "" for v in row]; break
        if ws: break
    if ws is None: raise SystemExit("No encontré la hoja con RUC/PENDIENTE en el módulo.")
    def c(*names):
        for n in names:
            for i,h in enumerate(HR):
                if norm(n)==h or norm(n) in h: return i
        return None
    iruc=c("CEDULA/RUC","RUC","CEDULA"); inom=c("CLIENTE","PROVEEDOR","NOMBRE","RAZON SOCIAL")
    ifec=c("FECHA"); ipen=c("PENDIENTE")
    pend={}; fechas={}
    for row in ws.iter_rows(min_row=hdr+1,values_only=True):
        if iruc is None or len(row)<=iruc: continue
        ruc=row[iruc]
        if ruc in (None,""): continue
        ruc=str(ruc).strip()
        p=row[ipen] if ipen is not None and len(row)>ipen else None
        try: p=float(p)
        except: continue
        f=pdate(row[ifec]) if ifec is not None and len(row)>ifec else None
        nom=str(row[inom]).strip() if inom is not None and len(row)>inom else ""
        pend.setdefault(ruc,{"nombre":nom,"total":0.0,"pre":0.0,"per":0.0})
        pend[ruc]["total"]+=p
    # segundo pase para split por anio se hace afuera con la fecha; reabrimos por simplicidad
    return ws, hdr, iruc, inom, ifec, ipen

def modulo_por_ruc(path, anio):
    import openpyxl
    wb=openpyxl.load_workbook(path,data_only=True)
    target=None; HR=None; hdr=None
    for sh in wb.worksheets:
        for i,row in enumerate(sh.iter_rows(values_only=True),1):
            vals=[norm(v) for v in row if v is not None]
            if any(("RUC" in v or "CEDULA" in v) for v in vals) and any("PENDIENTE" in v for v in vals):
                target=sh; hdr=i; HR=[norm(v) if v is not None else "" for v in row]; break
        if target: break
    if target is None: raise SystemExit("Módulo: no hallé hoja con RUC/PENDIENTE")
    def c(*names):
        for n in names:
            for i,h in enumerate(HR):
                if norm(n)==h or norm(n) in h: return i
        return None
    iruc=c("CEDULA/RUC","RUC","CEDULA"); inom=c("CLIENTE","PROVEEDOR","NOMBRE","RAZON SOCIAL")
    ifec=c("FECHA"); ipen=c("PENDIENTE")
    by={}
    for row in target.iter_rows(min_row=hdr+1,values_only=True):
        if len(row)<=iruc or row[iruc] in (None,""): continue
        ruc=str(row[iruc]).strip()
        try: p=float(row[ipen])
        except: continue
        f=pdate(row[ifec]) if ifec is not None else None
        nom=str(row[inom]).strip() if inom is not None and len(row)>inom else ""
        e=by.setdefault(ruc,{"nombre":nom,"total":0.0,"pre":0.0,"per":0.0})
        e["total"]+=p
        if f and anio and f.year>=anio: e["per"]+=p
        else: e["pre"]+=p
    for e in by.values():
        for k in ("total","pre","per"): e[k]=round(e[k],2)
    return by

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--mayor",required=True); ap.add_argument("--cuenta",required=True)
    ap.add_argument("--modulo",required=True); ap.add_argument("--apertura")
    ap.add_argument("--anio",type=int); ap.add_argument("--detalle")
    a=ap.parse_args()
    apertura_c, periodo_c, mov_c = cargar_mayor_cuenta(a.mayor,a.cuenta)
    inv = -1 if str(a.cuenta)[:1] in ("2","3","4") else 1
    apertura_c=round(apertura_c*inv,2); mov_c=round(mov_c*inv,2)
    periodo_c={k:round(v*inv,2) for k,v in periodo_c.items()}
    saldo_c=round(apertura_c+mov_c,2)
    anio=a.anio
    mod=modulo_por_ruc(a.modulo,anio or 0)
    pend_total=round(sum(e["total"] for e in mod.values()),2)
    pend_pre=round(sum(e["pre"] for e in mod.values()),2)
    pend_per=round(sum(e["per"] for e in mod.values()),2)
    D=round(saldo_c-pend_total,2)
    hist=round(apertura_c-pend_pre,2)
    per=round(mov_c-pend_per,2)
    print(f"CUENTA {a.cuenta}  (año corte: {anio or 's/d'})")
    print(f"  Saldo contable (mayor)      = {saldo_c:>14,.2f}  (apertura lump {apertura_c:,.2f} + período {mov_c:,.2f})")
    print(f"  Pendiente módulo            = {pend_total:>14,.2f}  (pre-año {pend_pre:,.2f} + año {pend_per:,.2f})")
    print(f"  DESCUADRE TOTAL D           = {D:>14,.2f}")
    print(f"     · Histórico (≤año-1)     = {hist:>14,.2f}")
    print(f"     · Período (año corte)    = {per:>14,.2f}")
    # Período por tercero: cruce mayor(beneficiario) vs módulo año (por nombre normalizado)
    mod_per_nom={}
    for ruc,e in mod.items():
        if e["per"]!=0: mod_per_nom[canon(e["nombre"])]={"ruc":ruc,"pend":e["per"]}
    nombres=set(periodo_c)|set(mod_per_nom)
    filas=[]
    for nom in nombres:
        cont=periodo_c.get(nom,0.0); modp=mod_per_nom.get(nom,{}).get("pend",0.0)
        dif=round(cont-modp,2)
        if abs(dif)>0.01: filas.append((nom,cont,modp,dif))
    filas.sort(key=lambda x:-abs(x[3]))
    print(f"\n  PERÍODO — terceros descuadrados ({len(filas)}):  (suma {round(sum(f[3] for f in filas),2):,.2f})")
    for nom,cont,modp,dif in filas[:25]:
        print(f"    {nom[:38]:<38} contab {cont:>11,.2f}  módulo {modp:>10,.2f}  dif {dif:>11,.2f}")
    if len(filas)>25: print(f"    ... y {len(filas)-25} más (ver --detalle)")
    # Histórico por tercero (si hay apertura del cierre anterior)
    if a.apertura:
        ape=modulo_por_ruc(a.apertura,anio or 0)  # su 'total' = pendiente al cierre anterior
        ape_total=round(sum(e["total"] for e in ape.values()),2)
        print(f"\n  APERTURA (cierre anterior) total módulo = {ape_total:,.2f}  vs lump contable {apertura_c:,.2f}  (dif {round(apertura_c-ape_total,2):,.2f})")
        rucs=set(ape)|set(mod)
        hfilas=[]
        for ruc in rucs:
            ape_t=ape.get(ruc,{}).get("total",0.0)         # pendiente al cierre anterior
            pre_corte=mod.get(ruc,{}).get("pre",0.0)        # pre-año aún pendiente al corte
            dif=round(ape_t-pre_corte,2)
            if abs(dif)>0.01:
                nom=ape.get(ruc,{}).get("nombre") or mod.get(ruc,{}).get("nombre","")
                hfilas.append((nom,ruc,ape_t,pre_corte,dif))
        hfilas.sort(key=lambda x:-abs(x[4]))
        print(f"  HISTÓRICO — terceros descuadrados ({len(hfilas)}):  (suma {round(sum(f[4] for f in hfilas),2):,.2f})")
        for nom,ruc,ape_t,pre_c,dif in hfilas[:25]:
            print(f"    {nom[:30]:<30} {ruc:<14} cierre {ape_t:>10,.2f}  pend.corte {pre_c:>10,.2f}  dif {dif:>10,.2f}")
    else:
        print(f"\n  (Sin --apertura no se abre el histórico por tercero; pasá el módulo al cierre del año anterior.)")

if __name__=="__main__":
    main()
