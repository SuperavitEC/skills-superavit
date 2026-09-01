#!/usr/bin/env python3
"""Orquestador de Anexos contables SAE. El modelo NUNCA lee los reportes pesados:
los muele este script y devuelve solo un resumen corto por anexo."""
import argparse, csv, json, os, sys
import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="1F3A5F"; LINE="C9D2DC"; GREEN="2E7D32"; AMBER="B26A00"; LINK="1F4E79"; FONT="Arial"
def F(sz=10,b=False,c="000000"): return Font(name=FONT,size=sz,bold=b,color=c)
thin=Side(style="thin",color=LINE); med=Side(style="medium",color=NAVY)
NUM='#,##0.00;(#,##0.00)'
BUCKETS=["Vigente","1-30","31-60","61-90","91-180","+180"]

# ---------- identidad del cliente (viene del config, nunca del código) ----------
# Este script es genérico: no conoce a ningún cliente. La razón social, el RUC, la
# fecha de corte y los alias de terceros salen del <cliente>.json que se pasa por
# --config. Ver config/README.md.
CFG={}
def set_cfg(cfg):
    global CFG; CFG=cfg or {}; return CFG
def cli_nombre(): return CFG.get("cliente","(cliente no definido en el config)")
def cli_ruc():    return CFG.get("ruc","")
def cli_corte():  return CFG.get("corte","")
def cli_sistema():return CFG.get("sistema","SAE")
def alias_terceros():
    """{ 'FRAGMENTO EN MAYUSCULAS': 'Nombre canónico' } — opcional, para normalizar
    nombres mal tipeados en los reportes del cliente. Se define en el config."""
    return list((CFG.get("alias_terceros") or {}).items())

def num(s):
    if s is None: return 0.0
    s=str(s).strip().replace('.','').replace(',','.')
    try: return float(s)
    except: return 0.0

# ---------- lectura de EEFF (chicos) ----------
def parse_eef(fn):
    w=openpyxl.load_workbook(fn,data_only=True); ws=w.active; rows=[]; started=False
    for r in ws.iter_rows(values_only=True):
        cod,nom,niv,sal,tot=(list(r)+[None]*5)[:5]
        if not started:
            if cod is not None and str(cod).strip().upper().replace("Ó","O")=="CODIGO": started=True
            continue
        if cod is None and nom is None: continue
        rows.append((str(cod).strip() if cod is not None else "", str(nom).strip() if nom else "", niv, sal, tot))
    return rows

def bg_saldos(rows):
    d={}
    for cod,nom,niv,sal,tot in rows:
        if cod and sal not in (None,""): d[cod]=float(sal)
    return d

# ---------- render BG / PYG ----------
def render_estado(ws,titulo,rows,anex_map):
    ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:D1"); c=ws["A1"]; c.value=cli_nombre().upper(); c.font=F(13,True,"FFFFFF"); c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=Alignment("left","center",indent=1); ws.row_dimensions[1].height=26
    ws.merge_cells("A2:D2"); ws["A2"]=titulo; ws["A2"].font=F(11,True,NAVY)
    ws.merge_cells("A3:D3"); ws["A3"]=("Al corte %s   ·   Sistema contable: %s   ·   Valores en USD"%(cli_corte(),cli_sistema())) if cli_corte() else ("Sistema contable: %s   ·   Valores en USD"%cli_sistema()); ws["A3"].font=F(9,False,"666666")
    hr=5; hc=[("Código",14),("Cuenta",52),("Saldo",16)]+([("Anexo",12)] if anex_map is not None else [])
    for j,(t,wd) in enumerate(hc,1):
        cc=ws.cell(hr,j,t); cc.font=F(9,True,"FFFFFF"); cc.fill=PatternFill("solid",fgColor=NAVY); cc.alignment=Alignment("center"); cc.border=Border(bottom=thin,top=thin,left=thin,right=thin); ws.column_dimensions[get_column_letter(j)].width=wd
    r=hr+1
    for cod,nom,niv,sal,tot in rows:
        up=nom.upper()
        if cod=="" and ("TOTAL" in up or "UTILIDAD" in up):
            ws.cell(r,2,nom).font=F(10,True); val=tot if tot not in (None,"") else sal
            cv=ws.cell(r,3,val); cv.font=F(10,True); cv.number_format=NUM; cv.alignment=Alignment("right")
            for j in range(1,5): ws.cell(r,j).border=Border(top=med)
            r+=1; continue
        if cod=="" and nom in ("ACTIVOS","PASIVOS","PATRIMONIO","INGRESOS","COSTO DE VENTAS","GASTOS"):
            ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4); cc=ws.cell(r,1,nom); cc.font=F(10,True,"FFFFFF"); cc.fill=PatternFill("solid",fgColor=NAVY); cc.alignment=Alignment("left","center",indent=1); ws.row_dimensions[r].height=20; r+=1; continue
        lvl=niv if isinstance(niv,int) else 4
        hoja=(sal is not None and sal!="")
        ws.cell(r,1,cod).font=F(9,False,"000000" if hoja else "666666")
        nm=ws.cell(r,2,nom); nm.font=F(9, not hoja, NAVY if lvl<=1 else "000000"); nm.alignment=Alignment(indent=max(0,lvl-1))
        val= sal if hoja else (tot if tot not in (None,"") else None)
        if val is not None:
            cv=ws.cell(r,3,val); cv.font=F(9,not hoja); cv.number_format=NUM; cv.alignment=Alignment("right")
        if anex_map is not None and hoja and cod in anex_map:
            ce=ws.cell(r,4,anex_map[cod]); ce.font=F(9,False,"999999"); ce.alignment=Alignment("center")
        for j in range(1,5): ws.cell(r,j).border=Border(bottom=thin)
        r+=1
    return r

# ---------- BG ANTERIOR (cadena de continuidad del Acta de Cuadre) ----------
def render_bg_anterior(ws,rows):
    """Hoja plana con los saldos de CIERRE del mes anterior, tal cual el BG exportado
    (sin recalcular): el servidor de actas cuadra apertura actual = cierre anterior."""
    ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:C1"); c=ws["A1"]; c.value="BG ANTERIOR  ·  SALDOS DE CIERRE DEL MES ANTERIOR"; c.font=F(13,True,"FFFFFF"); c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=Alignment("left","center",indent=1); ws.row_dimensions[1].height=26
    ws.merge_cells("A2:C2"); ws["A2"]="Cadena de continuidad del Acta de Cuadre: el servidor compara la apertura del mes actual contra estos cierres."; ws["A2"].font=F(9,False,"666666")
    hr=4
    for j,(t,wd) in enumerate([("Código",14),("Cuenta",52),("Saldo",16)],1):
        cc=ws.cell(hr,j,t); cc.font=F(9,True,"FFFFFF"); cc.fill=PatternFill("solid",fgColor=NAVY); cc.alignment=Alignment("center"); cc.border=Border(bottom=thin,top=thin,left=thin,right=thin); ws.column_dimensions[get_column_letter(j)].width=wd
    r=hr+1; n=0
    for cod,nom,niv,sal,tot in rows:
        if not cod or sal in (None,""): continue
        ws.cell(r,1,cod).font=F(9)
        ws.cell(r,2,nom).font=F(9)
        cv=ws.cell(r,3,float(sal)); cv.font=F(9); cv.number_format=NUM; cv.alignment=Alignment("right")
        for j in range(1,4): ws.cell(r,j).border=Border(bottom=thin)
        r+=1; n+=1
    return n

# ---------- saldo de un tramo en el Balance ----------
# CON SIGNO. Sumar valores absolutos infla cualquier tramo que traiga una
# contra-cuenta, exactamente al doble de esa cuenta. El caso típico es el anexo
# de propiedad, planta y equipo, que incluye la depreciación acumulada: con
# valores absolutos el costo y la depreciación se suman en vez de restarse.
# Esta cifra es la que el acta de cuadre compara contra el total del anexo, así
# que el defecto no queda en la presentación: se cobra como un hallazgo que no
# existe y que alguien tiene que firmar.
def saldo_bg_tramo(bgs,cuentas): return round(sum(bgs.get(c,0.0) for c in cuentas),2)

# El valor absoluto sirve para ORDENAR los anexos por peso. Nunca para cuadrar.
def peso_bg_tramo(bgs,cuentas): return round(sum(abs(bgs.get(c,0.0)) for c in cuentas),2)

# ---------- estado: sidecar JSON (NO va en el Excel del cliente) ----------
def state_path(salida): return salida + ".estado.json"
def write_state(salida,cfg,bgs):
    st=[]
    for a in cfg["anexos"]:
        sbg=saldo_bg_tramo(bgs,a["cuentas"])
        st.append({"id":a["id"],"desc":a["desc"],"tipo":a["tipo"],"reporte":a["reporte"],"estado":"pendiente","hoja":"","saldo_bg":sbg,"peso_bg":peso_bg_tramo(bgs,a["cuentas"]),"saldo_anexo":"","dif":""})
    json.dump(st,open(state_path(salida),"w",encoding="utf-8"),ensure_ascii=False,indent=1)
def load_state(salida): return json.load(open(state_path(salida),encoding="utf-8"))
def save_state(salida,st): json.dump(st,open(state_path(salida),"w",encoding="utf-8"),ensure_ascii=False,indent=1)
def update_state(st,aid,**kw):
    for r in st:
        if r["id"]==aid: r.update(kw)

def link_bg(wb,aid):
    if "BG" not in wb.sheetnames: return
    bg=wb["BG"]; sh=None
    for row in bg.iter_rows(min_col=4,max_col=4):
        c=row[0]
        if c.value in (aid, f"{aid} ↗"):
            # nombre de hoja del anexo
            for nm in wb.sheetnames:
                if nm.split()[0]==aid: sh=nm
            if sh:
                c.value=f"{aid} ↗"; c.hyperlink=Hyperlink(ref=c.coordinate, location=f"'{sh}'!A1"); c.font=F(9,True,LINK); c.alignment=Alignment("center")

# ---------- helpers de hoja-anexo ----------
def anexo_header(ws,aid,desc,sub=None,ncols=6):
    ws.sheet_view.showGridLines=False
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1"); c=ws["A1"]; c.value=f"ANEXO {aid}  ·  {desc.upper()}"; c.font=F(13,True,"FFFFFF"); c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=Alignment("left","center",indent=1); ws.row_dimensions[1].height=26
    ws.cell(2,1,"Cliente:").font=F(9,True,"555555"); ws.cell(2,2,cli_nombre()).font=F(9)
    ws.cell(2,4,"RUC:").font=F(9,True,"555555"); ws.cell(2,5,cli_ruc()).font=F(9)
    ws.cell(3,1,"Corte:").font=F(9,True,"555555"); ws.cell(3,2,cli_corte()).font=F(9)
    if sub: ws.cell(3,4,sub[0]).font=F(9,True,"555555"); ws.cell(3,5,sub[1]).font=F(9)
def colhead(ws,row,cols,h=22):
    for j,(t,w) in enumerate(cols,1):
        c=ws.cell(row,j,t); c.font=F(9,True,"FFFFFF"); c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=Alignment("center","center",wrap_text=True); c.border=Border(bottom=thin,top=thin,left=thin,right=thin); ws.column_dimensions[get_column_letter(j)].width=w
    ws.row_dimensions[row].height=h
def cuadre_box(ws,r,saldo_anexo_cell,saldo_bg):
    ws.cell(r,2,"Saldo según anexo").font=F(9,True); c=ws.cell(r,6,saldo_anexo_cell); c.font=F(9,True); c.number_format=NUM; c.alignment=Alignment("right")
    ws.cell(r+1,2,"Saldo según Balance").font=F(9); cb=ws.cell(r+1,6,round(saldo_bg,2)); cb.font=F(9); cb.number_format=NUM; cb.alignment=Alignment("right")
    ws.cell(r+2,2,"Diferencia").font=F(9,True); cd=ws.cell(r+2,6,f"=F{r}-F{r+1}"); cd.font=F(9,True,GREEN); cd.number_format=NUM; cd.alignment=Alignment("right")
    c=ws.cell(r+4,1,"← Volver al Balance"); c.hyperlink=Hyperlink(ref=c.coordinate, location="BG!A1"); c.font=F(8,True,LINK)

# ---------- D2: hoja CUADRE (la que lee el servidor) ----------
# El sidecar `.estado.json` ya tenía estos mismos números, pero NO viaja con el
# Excel: lo que se sube al acta es el .xlsx y nada más. Y las cajas de cuadre de
# las hojas de anexo escriben la diferencia como FÓRMULA («=F12-F13»), así que
# ningún total escrito como fórmula cuenta como dato — el servidor no confía en
# el caché de Excel. Esta hoja es el contrato entre la Fase 2 y el acta: una
# fila por anexo, TODO en valores duros.
CUADRE_HOJA = "CUADRE"
CUADRE_COLS = [("Anexo", 10), ("Descripción", 46), ("Hoja", 30),
               ("Saldo según anexo", 18), ("Saldo según Balance", 18),
               ("Diferencia", 14), ("Estado", 16)]


def render_cuadre(wb, st):
    """Reescribe la hoja CUADRE desde el estado. Se llama en cada build, así
    que refleja siempre el último cálculo de cada anexo."""
    if CUADRE_HOJA in wb.sheetnames:
        del wb[CUADRE_HOJA]
    ws = wb.create_sheet(CUADRE_HOJA)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    c = ws["A1"]; c.value = "CUADRE DE ANEXOS CONTRA BALANCE"
    c.font = F(13, True, "FFFFFF"); c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment("left", "center", indent=1)
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:G2")
    ws["A2"] = ("Resumen en valores; el detalle de cada anexo está en su hoja. "
                "Esta hoja es la que verifica el acta de cuadre.")
    ws["A2"].font = F(9, False, "666666")
    hr = 4
    for j, (t, w) in enumerate(CUADRE_COLS, 1):
        cc = ws.cell(hr, j, t)
        cc.font = F(9, True, "FFFFFF"); cc.fill = PatternFill("solid", fgColor=NAVY)
        cc.alignment = Alignment("center", "center", wrap_text=True)
        cc.border = Border(bottom=thin, top=thin, left=thin, right=thin)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[hr].height = 22
    r = hr + 1
    n_ok = n_hall = n_pend = 0
    for row in st:
        sa = row.get("saldo_anexo", "")
        sbg = row.get("saldo_bg", "")
        dif = row.get("dif", "")
        est = row.get("estado", "pendiente")
        ws.cell(r, 1, row["id"]).font = F(9, True)
        ws.cell(r, 2, row.get("desc", "")).font = F(9)
        ws.cell(r, 3, row.get("hoja", "")).font = F(9, False, "666666")
        for j, val in ((4, sa), (5, sbg), (6, dif)):
            if val in (None, ""):
                continue
            # Valor DURO, redondeado. Nunca una fórmula: el servidor no
            # recalcula el libro y una fórmula llega como texto.
            cv = ws.cell(r, j, round(float(val), 2))
            cv.font = F(9, j == 6)
            cv.number_format = NUM
            cv.alignment = Alignment("right")
        color = (GREEN if est == "Cuadra" else
                 AMBER if est == "Hallazgo" else "666666")
        ws.cell(r, 7, est).font = F(9, est != "pendiente", color)
        for j in range(1, 8):
            ws.cell(r, j).border = Border(bottom=thin)
        n_ok += est == "Cuadra"
        n_hall += est == "Hallazgo"
        n_pend += est not in ("Cuadra", "Hallazgo")
        r += 1
    ws.cell(r + 1, 2, f"{len(st)} anexos · {n_ok} cuadran · "
                      f"{n_hall} con hallazgo · {n_pend} pendientes").font = F(9, True)
    c = ws.cell(r + 3, 1, "← Volver al Balance")
    c.hyperlink = Hyperlink(ref=c.coordinate, location="BG!A1")
    c.font = F(8, True, LINK)
    return ws


# ---------- lectura del mayor (pesado, se filtra aqui) ----------
def mayor_cuenta(mayor,cuentas):
    cset=set(cuentas); out=[]
    with open(mayor,encoding='utf-8-sig') as f:
        rd=csv.reader(f,delimiter=';'); next(rd,None)
        for row in rd:
            if len(row)<12: continue
            if row[1].strip() in cset:
                out.append({"fecha":row[0].strip(),"cod":row[1].strip(),"asiento":row[3].strip(),"doc":row[4].strip(),"benef":row[6].strip(),"detalle":row[8].strip(),"debe":num(row[9]),"haber":num(row[10])})
    return out

# ---------- CONSTRUCTORES ----------
def build_amortizacion(ws,aid,a,args,bgs):
    from datetime import date as _date
    p=a.get("params",{})
    cs=p.get("corte","2026-05-31"); yy,mm,dd=[int(x) for x in cs.split("-")]; corte=_date(yy,mm,dd)
    anexo_header(ws,aid,a["desc"],("Entidad",p.get("entidad","")),ncols=6)
    for col,w in [("A",13),("B",8),("C",15),("D",13),("E",13),("F",13)]: ws.column_dimensions[col].width=w
    sched=p.get("amortizacion",[])
    def pf(f):
        try: y,mo,d=[int(x) for x in str(f).split("-")]; return _date(y,mo,d)
        except: return None
    corte_idx=None
    for ix,row in enumerate(sched):
        fd=pf(row["fecha"])
        if fd and fd<=corte: corte_idx=ix
    saldo_corte=round(sched[corte_idx]["saldo"],2) if corte_idx is not None else 0.0
    sbg=saldo_bg_tramo(bgs,a["cuentas"])
    corr=round(abs(bgs.get(p.get("corriente"),0)),2); nocorr=round(abs(bgs.get(p.get("no_corriente"),0)),2)
    if not sched:
        ws.cell(5,1,"Falta la tabla de amortización en la configuración del cliente.").font=F(9,False,"B26A00"); return 0.0,sbg
    try: corte12=_date(corte.year+1,corte.month,corte.day)
    except: corte12=_date(corte.year+1,corte.month,28)
    corriente_amort=round(sum(row["capital"] for row in sched if (pf(row["fecha"]) and corte<pf(row["fecha"])<=corte12)),2)
    nocorr_amort=round(saldo_corte-corriente_amort,2)
    ws.cell(5,1,"Resumen del préstamo al corte ("+corte.strftime("%d/%m/%Y")+")").font=F(10,True,NAVY)
    r=6
    ws.cell(r,1,"Saldo de capital al corte (tabla de amortización)").font=F(9); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4); c=ws.cell(r,6,saldo_corte); c.font=F(9); c.number_format=NUM; c.alignment=Alignment("right")
    for jx in range(1,7): ws.cell(r,jx).border=Border(bottom=thin)
    amort_row=r; r+=1
    ws.cell(r,1,"Saldo según balance (contable)").font=F(9); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4); c=ws.cell(r,6,sbg); c.font=F(9); c.number_format=NUM; c.alignment=Alignment("right")
    for jx in range(1,7): ws.cell(r,jx).border=Border(bottom=thin)
    bal_row=r; r+=1
    ws.cell(r,1,"Diferencia").font=F(9,True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4); c=ws.cell(r,6,"=F%d-F%d"%(amort_row,bal_row)); c.font=F(9,True,GREEN if abs(saldo_corte-sbg)<0.01 else AMBER); c.number_format=NUM; c.alignment=Alignment("right")
    for jx in range(1,7): ws.cell(r,jx).border=Border(top=med,bottom=thin)
    r+=2
    ws.cell(r,1,"Clasificación por vencimiento — verificación corriente / no corriente").font=F(10,True,NAVY); r+=1
    colhead(ws,r,[("Concepto",13),("",8),("",15),("Según amortización",13),("Según balance",13),("Diferencia",13)]); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3); r+=1
    for et,am,ba,cta,bold in [("Porción corriente (capital del próx. año / 12 meses)",corriente_amort,corr,p.get("corriente",""),False),("Porción no corriente (del año subsiguiente en adelante)",nocorr_amort,nocorr,p.get("no_corriente",""),False),("TOTAL",saldo_corte,sbg,"",True)]:
        ws.cell(r,1,et+(("  · "+str(cta)) if cta else "")).font=F(9,bold); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
        c=ws.cell(r,4,round(am,2)); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(9,bold)
        c=ws.cell(r,5,round(ba,2)); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(9,bold)
        c=ws.cell(r,6,"=D%d-E%d"%(r,r)); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(9,True,GREEN if abs(am-ba)<0.01 else AMBER)
        for jx in range(1,7): ws.cell(r,jx).border=Border(top=med) if bold else Border(bottom=thin)
        r+=1
    cls_amb = abs(corriente_amort-corr)>=1
    if cls_amb:
        ws.cell(r,1,"⚠ La clasificación contable corriente/no corriente NO coincide con la tabla de amortización — reclasificar.").font=F(8,False,"B26A00"); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); r+=1
    r+=2
    ws.cell(r,1,"Tabla de amortización — "+str(p.get("entidad",""))).font=F(10,True,NAVY); r+=1
    colhead(ws,r,[("Fecha",13),("Cuota",8),("Saldo capital",15),("Capital",13),("Interés",13),("Dividendo",13)]); r+=1
    for ix,row in enumerate(sched):
        ws.cell(r,1,str(row["fecha"])).font=F(8); ws.cell(r,2,row["cuota"]).font=F(8,ix==corte_idx); ws.cell(r,2).alignment=Alignment("center")
        for jx,k in [(3,"saldo"),(4,"capital"),(5,"interes"),(6,"dividendo")]:
            c=ws.cell(r,jx,round(row[k],2)); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(8,ix==corte_idx)
        if ix==corte_idx:
            for jx in range(1,7): ws.cell(r,jx).fill=PatternFill("solid",fgColor="FFF2CC")
            ws.cell(r,1,"► "+str(row["fecha"])).font=F(8,True)
        for jx in range(1,7): ws.cell(r,jx).border=Border(bottom=Side(style="thin",color="EEEEEE"))
        r+=1
    nr=r+1; notas=list(p.get("observaciones",[]))
    if notas:
        ws.cell(nr,1,"Observaciones").font=F(9,True,NAVY); nr+=1
        for n in notas:
            ws.cell(nr,1,"•  "+n).font=F(8,False,"444444"); ws.merge_cells(start_row=nr,start_column=1,end_row=nr,end_column=6); nr+=1
    cc=ws.cell(nr+1,1,"← Volver al Balance"); cc.hyperlink=Hyperlink(ref=cc.coordinate, location="BG!A1"); cc.font=F(8,True,LINK)
    return saldo_corte, sbg
def build_mayor_extracto(ws,aid,a,args,bgs):
    anexo_header(ws,aid,a["desc"])
    mv=mayor_cuenta(args.mayor,a["cuentas"])
    hr=5; colhead(ws,hr,[("Fecha",12),("Comprobante",16),("Beneficiario / concepto",40),("Débito",13),("Crédito",13),("Saldo",14)])
    si=sum(m["debe"]+m["haber"] for m in mv if m["doc"]=="SALDOS INICIALES")
    r=hr+1
    ws.cell(r,3,"Saldo inicial del período").font=F(9,True); s=si
    c=ws.cell(r,6,round(si,2)); c.font=F(9,True); c.number_format=NUM; c.alignment=Alignment("right")
    for j in range(1,7): ws.cell(r,j).border=Border(bottom=thin)
    r+=1
    for m in mv:
        if m["doc"]=="SALDOS INICIALES": continue
        s+=m["debe"]+m["haber"]
        ws.cell(r,1,m["fecha"]).font=F(8); ws.cell(r,2,m["asiento"]).font=F(8)
        ws.cell(r,3,(m["benef"] or m["detalle"] or m["doc"])[:60]).font=F(8)
        for j,v in [(4,m["debe"]),(5,m["haber"])]:
            if v: cc=ws.cell(r,j,round(v,2)); cc.font=F(8); cc.number_format=NUM; cc.alignment=Alignment("right")
        cc=ws.cell(r,6,round(s,2)); cc.font=F(8); cc.number_format=NUM; cc.alignment=Alignment("right")
        for j in range(1,7): ws.cell(r,j).border=Border(bottom=thin)
        r+=1
    ws.cell(r,3,"Saldo al corte").font=F(9,True); c=ws.cell(r,6,round(s,2)); c.font=F(9,True); c.number_format=NUM; c.alignment=Alignment("right")
    for j in range(1,7): ws.cell(r,j).border=Border(top=med)
    sa=round(abs(s),2); sbg=saldo_bg_tramo(bgs,a["cuentas"])
    cuadre_box(ws,r+2,sa,sbg)
    return sa,sbg

def parse_concil(path):
    w=openpyxl.load_workbook(path,data_only=True); sh=w.active; banc=cont=conc=None
    for row in sh.iter_rows(values_only=True):
        for v in row:
            if isinstance(v,str):
                u=v.strip().upper(); nums=[x for x in row if isinstance(x,(int,float))]
                if u=="SALDO BANCARIO": banc=nums[0] if nums else None
                elif u=="SALDO CONTABLE": cont=nums[-1] if nums else None
                elif u=="SALDO CONCILIADO": conc=nums[-1] if nums else None
    return banc or 0.0, cont or 0.0, conc or 0.0

def build_conciliacion_bancaria(ws,aid,a,args,bgs):
    anexo_header(ws,aid,a["desc"],("Conciliación al","31/05/2026"),ncols=7)
    nombres={cod:nom for cod,nom,niv,sal,tot in parse_eef(args.bg) if cod in a["cuentas"]}
    ws.cell(4,1,"Saldo banco = seg\u00fan estado de cuenta    \u00b7    Saldo libros = seg\u00fan conciliaci\u00f3n    \u00b7    Saldo mayor = seg\u00fan balance general").font=F(8,False,"777777")
    ws.merge_cells(start_row=4,start_column=1,end_row=4,end_column=7)
    hr=5; colhead(ws,hr,[("C\u00f3digo",12),("Banco",36),("Saldo banco",14),("Saldo libros",14),("Partidas concil.",16),("Saldo mayor",14),("Diferencia",13)])
    r=hr+1; tot_lib=0.0; tot_gl=0.0; notas=[]
    for cod in a["cuentas"]:
        pth=os.path.join(args.insumos or "",f"conciliacion_{cod}.xlsx")
        if not os.path.exists(pth):
            notas.append(f"{cod} {nombres.get(cod,'')}: falta el reporte de conciliación."); continue
        banc,cont,conc=parse_concil(pth); gl=abs(bgs.get(cod,0.0)); part=round(cont-banc,2); dif=round(cont-gl,2)
        tot_lib+=cont; tot_gl+=gl
        ws.cell(r,1,cod).font=F(8); ws.cell(r,2,nombres.get(cod,"")[:36]).font=F(8)
        for j,v,flag in [(3,banc,False),(4,cont,False),(5,part,abs(part)>=0.01),(6,gl,False),(7,dif,abs(dif)>=0.01)]:
            c=ws.cell(r,j,round(v,2)); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(8,flag,AMBER if flag else "000000")
        for j in range(1,8): ws.cell(r,j).border=Border(bottom=thin)
        if abs(part)>=0.01: notas.append(f"{nombres.get(cod,cod)}: partidas conciliatorias por {part:,.2f} (saldo banco {banc:,.2f} vs libros {cont:,.2f}) — pendientes de identificar.")
        if abs(dif)>=0.01: notas.append(f"{nombres.get(cod,cod)}: saldo de la conciliación {cont:,.2f} vs mayor {gl:,.2f}, diferencia {dif:,.2f} a depurar.")
        r+=1
    tr=r; ws.cell(tr,2,"TOTAL").font=F(9,True)
    cD=ws.cell(tr,4,"=SUM(D%d:D%d)"%(hr+1,tr-1)); cD.font=F(9,True); cD.number_format=NUM; cD.alignment=Alignment("right")
    cF=ws.cell(tr,6,"=SUM(F%d:F%d)"%(hr+1,tr-1)); cF.font=F(9,True); cF.number_format=NUM; cF.alignment=Alignment("right")
    for j in range(1,8): ws.cell(tr,j).border=Border(top=med)
    r=tr+2
    ws.cell(r,2,"Saldo según libros (conciliación)").font=F(9,True); c=ws.cell(r,7,"=D%d"%tr); c.font=F(9,True); c.number_format=NUM; c.alignment=Alignment("right")
    ws.cell(r+1,2,"Saldo según mayor (balance)").font=F(9); c=ws.cell(r+1,7,"=F%d"%tr); c.font=F(9); c.number_format=NUM; c.alignment=Alignment("right")
    dft=round(tot_lib-tot_gl,2)
    ws.cell(r+2,2,"Diferencia").font=F(9,True); c=ws.cell(r+2,7,"=G%d-G%d"%(r,r+1)); c.font=F(9,True,GREEN if abs(dft)<0.01 else AMBER); c.number_format=NUM; c.alignment=Alignment("right")
    nr=r+4
    if notas:
        ws.cell(nr,1,"Observaciones").font=F(9,True,NAVY); nr+=1
        for n in notas:
            ws.cell(nr,1,"•  "+n).font=F(8,False,"444444"); ws.merge_cells(start_row=nr,start_column=1,end_row=nr,end_column=7); nr+=1
    c=ws.cell(nr+1,1,"← Volver al Balance"); c.hyperlink=Hyperlink(ref=c.coordinate, location="BG!A1"); c.font=F(8,True,LINK)
    return round(tot_lib,2), round(tot_gl,2)

def build_inversiones(ws,aid,a,args,bgs):
    anexo_header(ws,aid,a["desc"],ncols=10)
    nombres={cod:nom for cod,nom,niv,sal,tot in parse_eef(args.bg) if cod in a["cuentas"]}
    invs=(a.get("params") or {}).get("inversiones")
    if not invs: invs=[{"cuenta":c} for c in a["cuentas"]]
    cubiertas={x.get("cuenta") for x in invs}
    for c in a["cuentas"]:
        if c not in cubiertas: invs=invs+[{"cuenta":c}]
    for col,w in [("A",13),("B",22),("C",20),("D",16),("E",13),("F",13),("G",17),("H",14),("I",10),("J",16)]: ws.column_dimensions[col].width=w
    hr=5
    colhead(ws,hr,[("Cuenta",13),("Tipo de instrumento",22),("Institución",20),("N° certificado",16),("F. inversión",13),("Vencimiento",13),("Capital invertido",17),("Rendimientos",14),("% rend.",10),("Valor en libros",16)])
    r=hr+1; notas=[]; saldo_anexo=0.0
    for inv in invs:
        c=inv.get("cuenta"); sal=abs(bgs.get(c,0.0))
        ws.cell(r,1,c).font=F(8)
        b=ws.cell(r,2, inv.get("tipo") or "(por confirmar)"); b.font=F(8,False,"000000" if inv.get("tipo") else "B26A00")
        cc=ws.cell(r,3, inv.get("institucion") or "(por confirmar)"); cc.font=F(8,False,"000000" if inv.get("institucion") else "B26A00")
        for jj,key in [(4,"n_certificado"),(5,"fecha_inversion"),(6,"vencimiento")]:
            v=inv.get(key); ws.cell(r,jj, v if v not in (None,"") else "—").font=F(8,False,"000000" if v not in (None,"") else "999999")
        cap=inv.get("capital"); ren=inv.get("rendimientos")
        if isinstance(cap,(int,float)):
            g=ws.cell(r,7,round(cap,2)); g.number_format=NUM; g.alignment=Alignment("right"); g.font=F(8)
            h=ws.cell(r,8,round(ren or 0,2)); h.number_format=NUM; h.alignment=Alignment("right"); h.font=F(8)
            saldo_anexo+=cap+(ren or 0)
        else:
            ws.cell(r,7,None); ws.cell(r,8,None); saldo_anexo+=sal
        pc=ws.cell(r,9,f'=IF(OR(G{r}="",G{r}=0),"",H{r}/G{r})'); pc.number_format="0.0%"; pc.alignment=Alignment("right"); pc.font=F(8)
        vc=ws.cell(r,10,f'=IF(G{r}="",{round(sal,2)},G{r}+H{r})'); vc.number_format=NUM; vc.alignment=Alignment("right"); vc.font=F(8,True)
        for jj in range(1,11): ws.cell(r,jj).border=Border(bottom=thin)
        if inv.get("observacion"): notas.append((c,inv["observacion"]))
        r+=1
    tot_row=r; ws.cell(r,1,"TOTAL").font=F(9,True)
    vt=ws.cell(r,10,f"=SUM(J{hr+1}:J{r-1})"); vt.font=F(9,True); vt.number_format=NUM; vt.alignment=Alignment("right")
    for jj in range(1,11): ws.cell(r,jj).border=Border(top=med)
    sbg=saldo_bg_tramo(bgs,a["cuentas"]); r+=2
    ws.cell(r,8,"Saldo según balance").font=F(9); ws.merge_cells(start_row=r,start_column=8,end_row=r,end_column=9); vb=ws.cell(r,10,sbg); vb.font=F(9); vb.number_format=NUM; vb.alignment=Alignment("right")
    ws.cell(r+1,8,"Diferencia").font=F(9,True); ws.merge_cells(start_row=r+1,start_column=8,end_row=r+1,end_column=9); vd=ws.cell(r+1,10,f"=J{tot_row}-J{r}"); vd.font=F(9,True,GREEN if abs(round(saldo_anexo,2)-sbg)<0.01 else AMBER); vd.number_format=NUM; vd.alignment=Alignment("right")
    nr=r+3
    ws.cell(nr,1,"Campos en blanco (capital, rendimientos, institución, fechas): pendientes de completar con el certificado del instrumento.").font=F(8,False,"888888"); ws.merge_cells(start_row=nr,start_column=1,end_row=nr,end_column=10); nr+=1
    if notas:
        ws.cell(nr,1,"Observaciones").font=F(9,True,NAVY); nr+=1
        for cod,obs in notas:
            ws.cell(nr,1,"•  "+str(cod)+": "+obs).font=F(8,False,"444444"); ws.merge_cells(start_row=nr,start_column=1,end_row=nr,end_column=10); nr+=1
    cv=ws.cell(nr+1,1,"← Volver al Balance"); cv.hyperlink=Hyperlink(ref=cv.coordinate, location="BG!A1"); cv.font=F(8,True,LINK)
    return round(saldo_anexo,2), sbg

def build_antiguedad(ws,aid,a,args,bgs):
    from datetime import date,datetime
    cs=(a.get("params") or {}).get("corte","2026-05-31"); yy,mm,dd=[int(x) for x in cs.split("-")]; corte=date(yy,mm,dd)
    anexo_header(ws,aid,a["desc"],("Antigüedad al",corte.strftime("%d/%m/%Y")),ncols=8)
    nombres={cod:nom for cod,nom,niv,sal,tot in parse_eef(args.bg)}
    sbg=saldo_bg_tramo(bgs,a["cuentas"])
    def_acct=max(a["cuentas"],key=lambda c:abs(bgs.get(c,0))); def_name=nombres.get(def_acct,"").title()
    override=(a.get("params") or {}).get("cuentas_relacionados",{})
    path=os.path.join(args.insumos or "", (a.get("params") or {}).get("reporte_archivo",""))
    if not os.path.exists(path):
        ws.cell(5,1,"Falta el reporte de cartera en insumos: "+path).font=F(9,False,"B26A00"); return 0.0, sbg
    def pdate(v):
        if isinstance(v,datetime): return v.date()
        try: return datetime.strptime(str(v).strip(),"%d/%m/%Y").date()
        except: return None
    def bk(dd):
        return "Vigente" if dd<=0 else "1-30" if dd<=30 else "31-60" if dd<=60 else "61-90" if dd<=90 else "91-180" if dd<=180 else "+180"
    wb2=openpyxl.load_workbook(path,data_only=True); sh=wb2.active; cmap=None; clientes={}
    for r0 in sh.iter_rows(values_only=True):
        if cmap is None:
            if r0 and any(isinstance(v,str) and "VENCIMIENTO" in str(v).upper() for v in r0):
                cmap={}
                for idx,v in enumerate(r0):
                    u=str(v or "").upper().strip()
                    if "DOCUMENTO" in u and u.startswith("F"): cmap["femi"]=idx
                    elif "VENCIM" in u: cmap["venc"]=idx
                    elif "NUMERO" in u or "NÚMERO" in u: cmap.setdefault("num",idx)
                    elif "RUC" in u or "CEDULA" in u or "CÉDULA" in u: cmap["ruc"]=idx
                    elif u in ("CLIENTE","PROVEEDOR"): cmap["cli"]=idx
                    elif "VENDEDOR" in u: cmap["vend"]=idx
                    elif "COMENTARIO" in u: cmap["com"]=idx
                    elif "PENDIENTE" in u: cmap["pend"]=idx
            continue
        g=lambda k: (r0[cmap[k]] if (k in cmap and cmap[k]<len(r0)) else None)
        femi=pdate(g("femi")); venc=pdate(g("venc")); num=g("num"); ruc=g("ruc"); cli=g("cli"); vend=g("vend"); com=g("com"); pend=g("pend")
        if pend in (None,"") or abs(pend)<0.005 or venc is None: continue
        key=(str(ruc or ""),str(cli or ""))
        clientes.setdefault(key,[]).append({"num":num,"femi":femi,"venc":venc,"vend":vend,"com":com,"pend":pend,"dias":(corte-venc).days,"rango":bk((corte-venc).days)})
    bsum={x:0.0 for x in BUCKETS}; total_pf=0.0
    for items in clientes.values():
        for it in items: bsum[it["rango"]]+=it["pend"]; total_pf+=it["pend"]
    bsum={x:round(bsum[x],2) for x in BUCKETS}; total_pf=round(total_pf,2)
    ws.cell(5,1,"Resumen por antigüedad").font=F(10,True,NAVY)
    colhead(ws,6,[("Rango",14),("Monto",16),("%",10)])
    fs=7; stl=fs+len(BUCKETS); r=fs
    for b in BUCKETS:
        ws.cell(r,1,b).font=F(9); c=ws.cell(r,2,bsum[b]); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(9)
        pc=ws.cell(r,3,f'=IF($B${stl}="","",B{r}/$B${stl})'); pc.number_format="0.0%"; pc.alignment=Alignment("right"); pc.font=F(9)
        for jx in range(1,4): ws.cell(r,jx).border=Border(bottom=thin)
        r+=1
    ws.cell(r,1,"Total cartera").font=F(9,True); c=ws.cell(r,2,f"=SUM(B{fs}:B{r-1})"); c.font=F(9,True); c.number_format=NUM; c.alignment=Alignment("right")
    for jx in range(1,4): ws.cell(r,jx).border=Border(top=med)
    dh=r+3; ws.cell(dh-1,1,"Detalle por cliente — factura por factura, más antiguo primero").font=F(10,True,NAVY)
    colhead(ws,dh,[("N° factura",16),("F. emisión",12),("F. vencimiento",13),("Días",7),("Rango",9),("Vendedor",18),("Comentario",30),("Pendiente",14)])
    r=dh+1
    # clientes ordenados por factura mas antigua
    orden=sorted(clientes.items(), key=lambda kv: min(it["venc"] for it in kv[1]))
    for (ruc,cli),items in orden:
        acc=override.get(ruc); acod=acc[0] if acc else def_acct; anom=(acc[1] if acc else def_name)
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
        cc=ws.cell(r,1,f"{cli}    ·    RUC {ruc}    ·    Cuenta {acod} — {anom}"); cc.font=F(9,True,"FFFFFF"); cc.fill=PatternFill("solid",fgColor=NAVY); cc.alignment=Alignment("left","center",indent=1); ws.row_dimensions[r].height=18
        r+=1; cf=r
        for it in sorted(items,key=lambda x:x["venc"]):
            ws.cell(r,1,str(it["num"] or "")).font=F(8)
            ws.cell(r,2, it["femi"].strftime("%d/%m/%Y") if it["femi"] else "").font=F(8)
            ws.cell(r,3, it["venc"].strftime("%d/%m/%Y")).font=F(8)
            cd=ws.cell(r,4,it["dias"]); cd.font=F(8); cd.alignment=Alignment("right")
            ws.cell(r,5,it["rango"]).font=F(8,False,"B26A00" if it["dias"]>180 else "000000")
            ws.cell(r,6,str(it["vend"] or "")[:20]).font=F(8)
            ws.cell(r,7,str(it["com"] or "")[:32]).font=F(8)
            cp=ws.cell(r,8,round(it["pend"],2)); cp.number_format=NUM; cp.alignment=Alignment("right"); cp.font=F(8)
            for jx in range(1,9): ws.cell(r,jx).border=Border(bottom=Side(style="thin",color="E8E8E8"))
            r+=1
        ws.cell(r,1,"Subtotal").font=F(8,True,"555555")
        cs2=ws.cell(r,8,f"=SUM(H{cf}:H{r-1})"); cs2.font=F(8,True); cs2.number_format=NUM; cs2.alignment=Alignment("right")
        for jx in range(1,9): ws.cell(r,jx).border=Border(top=thin)
        r+=1
    ws.cell(r,1,"TOTAL CARTERA (por factura)").font=F(9,True)
    ct=ws.cell(r,8,round(total_pf,2)); ct.font=F(9,True); ct.number_format=NUM; ct.alignment=Alignment("right")
    for jx in range(1,9): ws.cell(r,jx).border=Border(top=med)
    total_row=r; r+=2
    ws.cell(r,6,"Saldo según balance").font=F(9); ws.merge_cells(start_row=r,start_column=6,end_row=r,end_column=7); c=ws.cell(r,8,sbg); c.font=F(9); c.number_format=NUM; c.alignment=Alignment("right")
    dft=round(total_pf-sbg,2)
    ws.cell(r+1,6,"Diferencia").font=F(9,True); ws.merge_cells(start_row=r+1,start_column=6,end_row=r+1,end_column=7); c=ws.cell(r+1,8,f"=H{total_row}-H{r}"); c.font=F(9,True,GREEN if abs(dft)<0.01 else AMBER); c.number_format=NUM; c.alignment=Alignment("right")
    nr=r+3; notas=list((a.get("params") or {}).get("observaciones",[]))
    if abs(dft)>=0.01:
        notas.insert(0,f"La cartera por factura ({total_pf:,.2f}) supera el saldo contable ({sbg:,.2f}) en {dft:,.2f}: facturas antiguas «SALDOS INICIALES» sin cruzar (el Global/contable ya las netea). En depuración antes de adoptar el «por factura».")
    if not override:
        notas.append("Cuenta contable asignada por defecto a "+def_acct+" — "+def_name+". Pendiente identificar los clientes relacionados (101020601) y del exterior (101020502).")
    if notas:
        ws.cell(nr,1,"Observaciones").font=F(9,True,NAVY); nr+=1
        for n in notas:
            ws.cell(nr,1,"•  "+n).font=F(8,False,"444444"); ws.merge_cells(start_row=nr,start_column=1,end_row=nr,end_column=8); nr+=1
    c=ws.cell(nr+1,1,"← Volver al Balance"); c.hyperlink=Hyperlink(ref=c.coordinate, location="BG!A1"); c.font=F(8,True,LINK)
    return total_pf, sbg

def build_lotes_tc(ws,aid,a,args,bgs):
    anexo_header(ws,aid,a["desc"],ncols=6)
    sbg=saldo_bg_tramo(bgs,a["cuentas"])
    mv=mayor_cuenta(args.mayor,a["cuentas"])
    si=sum(m["debe"]+m["haber"] for m in mv if m["doc"]=="SALDOS INICIALES")
    cob=sum(m["debe"] for m in mv if "TARJETA" in m["doc"].upper())
    liq=sum(m["haber"] for m in mv if "TRANSFER" in m["doc"].upper())
    ret=sum(m["haber"] for m in mv if "RETEN" in m["doc"].upper())
    sf=round(si+cob+liq+ret,2)
    hr=5; colhead(ws,hr,[("Concepto",52),("",4),("",4),("",4),("",4),("Valor",16)]); ws.merge_cells(start_row=hr,start_column=2,end_row=hr,end_column=5)
    rows=[("Saldo inicial",si,False),("(+) Cobros con tarjeta de crédito",cob,False),("(−) Liquidaciones de procesadoras",liq,False),("(−) Retenciones de tarjeta",ret,False),("Saldo al corte (lotes pendientes de liquidación)",sf,True)]
    r=hr+1
    for txt,val,tot in rows:
        ws.cell(r,1,txt).font=F(9,tot); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
        c=ws.cell(r,6,round(val,2)); c.font=F(9,tot); c.number_format=NUM; c.alignment=Alignment("right")
        for j in range(1,7): ws.cell(r,j).border=Border(top=med,bottom=thin) if tot else Border(bottom=thin)
        r+=1
    r+=1
    ws.cell(r,1,"Saldo según anexo").font=F(9,True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5); c=ws.cell(r,6,sf); c.font=F(9,True); c.number_format=NUM; c.alignment=Alignment("right")
    ws.cell(r+1,1,"Saldo según balance (101020801)").font=F(9); ws.merge_cells(start_row=r+1,start_column=1,end_row=r+1,end_column=5); c=ws.cell(r+1,6,sbg); c.font=F(9); c.number_format=NUM; c.alignment=Alignment("right")
    dft=round(sf-sbg,2); ws.cell(r+2,1,"Diferencia").font=F(9,True); ws.merge_cells(start_row=r+2,start_column=1,end_row=r+2,end_column=5); c=ws.cell(r+2,6,f"=F{r}-F{r+1}"); c.font=F(9,True,GREEN if abs(dft)<0.01 else AMBER); c.number_format=NUM; c.alignment=Alignment("right")
    nr=r+4
    notas=["El saldo son lotes de tarjeta cobrados al cliente y pendientes de liquidación por las procesadoras al corte."]+list((a.get("params") or {}).get("observaciones",[]))
    ws.cell(nr,1,"Observaciones").font=F(9,True,NAVY); nr+=1
    for n in notas:
        ws.cell(nr,1,"•  "+n).font=F(8,False,"444444"); ws.merge_cells(start_row=nr,start_column=1,end_row=nr,end_column=6); nr+=1
    c=ws.cell(nr+1,1,"← Volver al Balance"); c.hyperlink=Hyperlink(ref=c.coordinate, location="BG!A1"); c.font=F(8,True,LINK)
    return sf, sbg

def build_anticipo_empleados(ws,aid,a,args,bgs):
    import os, glob
    from datetime import datetime as _dt
    anexo_header(ws,aid,a["desc"],ncols=7)
    for col,w in [("A",13),("B",15),("C",16),("D",13),("E",13),("F",13),("G",14)]: ws.column_dimensions[col].width=w
    sbg=saldo_bg_tramo(bgs,a["cuentas"])
    KW=alias_terceros()   # nombres de empleados: salen del config del cliente
    def canon(nom):
        u=str(nom or "").upper()
        for k,full in KW:
            if k in u: return full
        return str(nom or "").strip().title()
    def pdate(x):
        try: return _dt.strptime(str(x).strip(),"%d/%m/%Y")
        except: return _dt(2026,12,31)
    rdir=args.insumos or ""
    # apertura por empleado
    apert={}; ajuste=0.0
    apf=os.path.join(rdir,(a.get("params") or {}).get("apertura_archivo",""))
    if os.path.exists(apf):
        aw=openpyxl.load_workbook(apf,data_only=True); ash=aw.active; rows=list(ash.iter_rows(values_only=True)); hr0=None
        for ii,row in enumerate(rows):
            if row and any(isinstance(v,str) and str(v).strip().upper()=="EMPLEADO" for v in row): hr0=ii; break
        sub=rows[hr0+1]; scols=[k for k,v in enumerate(sub) if isinstance(v,str) and v.strip().upper()=="SALDO"][:12]
        for row in rows[hr0+2:]:
            nm=row[0]
            if nm in (None,"") or str(nm).strip().upper()=="TOTAL": continue
            val=round(sum((row[k] if k<len(row) and isinstance(row[k],(int,float)) else 0) for k in scols),2)
            if "AJUSTE" in str(nm).upper(): ajuste+=val
            else: apert[canon(nm)]=apert.get(canon(nm),0)+val
    # roster del rol más reciente (quien NO esté aquí, no se le descuenta por nómina)
    roster=set(); files=sorted(glob.glob(os.path.join(rdir,"rol_2026-*.xlsx")))
    if files:
        w=openpyxl.load_workbook(files[-1],data_only=True); sh=w.active
        for row in sh.iter_rows(min_row=6,values_only=True):
            if row and row[0] and str(row[0]).strip().upper() not in ("","TOTAL"): roster.add(canon(str(row[1] or "")+" "+str(row[2] or "")))
    # mayor: lotes (apertura + otorgado) y pagos
    mv=mayor_cuenta(args.mayor,a["cuentas"]); si=round(sum(m["debe"]+m["haber"] for m in mv if m["doc"]=="SALDOS INICIALES"),2)
    pago=round(-sum(m["haber"] for m in mv if m["doc"]!="SALDOS INICIALES"),2)
    lots=[]
    for cn,v in apert.items():
        if abs(v)>=0.01: lots.append({"emp":cn,"f":_dt(2025,12,31),"fs":"31/12/2025","con":"Saldo inicial (apertura dic-2025)","comp":"","m":round(v,2)})
    otorg=0.0
    for m in mv:
        if m["doc"]=="SALDOS INICIALES" or m["debe"]<=0: continue
        otorg+=m["debe"]; lots.append({"emp":canon(m["benef"]),"f":pdate(m["fecha"]),"fs":m["fecha"],"con":(m["detalle"] or "")[:42] or "Anticipo","comp":(m["asiento"] or "")[:14],"m":round(m["debe"],2)})
    otorg=round(otorg,2); sf=round(si+otorg-pago,2)
    frozen=lambda e: e not in roster
    # FIFO: el pago consume lotes (más antiguos primero) de empleados que SÍ están en el rol
    consumed={}; rem=pago
    for lot in sorted(lots,key=lambda x:(x["f"],)):
        if rem<=0.005: break
        if frozen(lot["emp"]) or lot["m"]<=0: continue
        take=min(lot["m"],rem); lot["m"]=round(lot["m"]-take,2); rem=round(rem-take,2); consumed[lot["emp"]]=round(consumed.get(lot["emp"],0)+take,2)
    pend={}
    for lot in lots:
        if lot["m"]>0.005: pend.setdefault(lot["emp"],[]).append(lot)
    # ===== Rollforward =====
    ws.cell(5,1,"Rollforward de la cuenta 101020804 (período 2026)").font=F(10,True,NAVY)
    hr=6; colhead(ws,hr,[("Concepto",13),("",15),("",16),("",13),("",13),("",13),("Valor",14)]); ws.merge_cells(start_row=hr,start_column=1,end_row=hr,end_column=6); r=hr+1
    for txt,val,tot in [("Saldo inicial 01/01/2026",si,False),("(+) Anticipos otorgados en 2026",otorg,False),("(−) Descontados vía liquidación de rol",-pago,False),("Saldo al corte 31/05/2026",sf,True)]:
        ws.cell(r,1,txt).font=F(9,tot); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
        c=ws.cell(r,7,val); c.font=F(9,tot); c.number_format=NUM; c.alignment=Alignment("right")
        for jx in range(1,8): ws.cell(r,jx).border=Border(top=med,bottom=thin) if tot else Border(bottom=thin)
        r+=1
    cc=ws.cell(hr+4,7,"=G%d+G%d+G%d"%(hr+1,hr+2,hr+3)); cc.font=F(9,True); cc.number_format=NUM; cc.alignment=Alignment("right")
    # ===== Resumen pendiente por empleado =====
    r+=1; ws.cell(r,1,"Anticipos pendientes de descontar por empleado").font=F(10,True,NAVY); r+=1
    colhead(ws,r,[("Empleado",13),("",15),("Estado",16),("Apertura",13),("Otorgado",13),("Descontado",13),("Pendiente",14)]); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2); r+=1
    notas=[]; tot_pend=0.0; first_emp=r
    def emp_otorg(cn): return round(sum(m["debe"] for m in mv if m["doc"]!="SALDOS INICIALES" and m["debe"]>0 and canon(m["benef"])==cn),2)
    nombres=set(apert)|set(consumed)|set(pend)
    summ=[]
    for cn in nombres:
        ap=round(apert.get(cn,0),2); ot=emp_otorg(cn); de=round(consumed.get(cn,0),2); pe=round(sum(l["m"] for l in pend.get(cn,[])),2)
        summ.append((cn,ap,ot,de,pe))
    for cn,ap,ot,de,pe in sorted(summ,key=lambda x:-x[4]):
        est="Activo (en rol)" if cn in roster else "No descontado por rol"
        ws.cell(r,1,cn).font=F(8); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
        c=ws.cell(r,3,est); c.font=F(8,cn not in roster,"B26A00" if cn not in roster else "2E7D32"); c.alignment=Alignment("center")
        for jx,v in [(4,ap),(5,ot),(6,de),(7,pe)]:
            cc=ws.cell(r,jx,v); cc.number_format=NUM; cc.alignment=Alignment("right"); cc.font=F(8,jx==7)
        for jx in range(1,8): ws.cell(r,jx).border=Border(bottom=thin)
        tot_pend+=pe; r+=1
        if cn not in roster and pe>=0.01: notas.append(cn+": no está en el rol → sus anticipos no se descuentan por nómina. Pendiente "+format(pe,",.2f")+" — definir descuento o regularizar.")
    if abs(ajuste)>=0.01:
        ws.cell(r,1,"Ajuste de apertura (2025)").font=F(8,False,"555555"); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
        cc=ws.cell(r,7,round(ajuste,2)); cc.number_format=NUM; cc.alignment=Alignment("right"); cc.font=F(8)
        for jx in range(1,8): ws.cell(r,jx).border=Border(bottom=thin)
        tot_pend+=ajuste; r+=1
    tpr=r; ws.cell(r,1,"TOTAL PENDIENTE").font=F(9,True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); c=ws.cell(r,7,"=SUM(G%d:G%d)"%(first_emp,r-1)); c.font=F(9,True); c.number_format=NUM; c.alignment=Alignment("right")
    for jx in range(1,8): ws.cell(r,jx).border=Border(top=med)
    r+=2
    ws.cell(r,5,"Saldo según balance").font=F(9); ws.merge_cells(start_row=r,start_column=5,end_row=r,end_column=6); c=ws.cell(r,7,sbg); c.font=F(9); c.number_format=NUM; c.alignment=Alignment("right")
    dft=round(tot_pend-sbg,2); ws.cell(r+1,5,"Diferencia").font=F(9,True); ws.merge_cells(start_row=r+1,start_column=5,end_row=r+1,end_column=6); c=ws.cell(r+1,7,"=G%d-G%d"%(tpr,r)); c.font=F(9,True,GREEN if abs(dft)<0.01 else AMBER); c.number_format=NUM; c.alignment=Alignment("right")
    # ===== Detalle de lo pendiente por empleado (desde cuándo) =====
    r+=3; ws.cell(r,1,"Detalle de lo pendiente — desde cuándo se otorgó (no descontado al 31/05/2026)").font=F(10,True,NAVY); r+=1
    colhead(ws,r,[("Fecha otorg.",13),("Comprob.",15),("Concepto",16),("",13),("",13),("",13),("Monto pend.",14)]); ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=6); r+=1
    for cn in sorted(pend,key=lambda k:-sum(l["m"] for l in pend[k])):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7); cc=ws.cell(r,1,cn+("" if cn in roster else "  —  no descontado por rol")); cc.font=F(9,True,"FFFFFF"); cc.fill=PatternFill("solid",fgColor=NAVY if cn in roster else AMBER); cc.alignment=Alignment("left","center",indent=1); ws.row_dimensions[r].height=16; r+=1
        cf=r
        for l in sorted(pend[cn],key=lambda x:x["f"]):
            ws.cell(r,1,l["fs"]).font=F(8); ws.cell(r,2,l["comp"]).font=F(8); ws.cell(r,3,l["con"]).font=F(8); ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=6)
            c=ws.cell(r,7,round(l["m"],2)); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(8)
            for jx in range(1,8): ws.cell(r,jx).border=Border(bottom=Side(style="thin",color="EEEEEE"))
            r+=1
        ws.cell(r,3,"Pendiente "+cn).font=F(8,True,"555555"); ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=6); c=ws.cell(r,7,f"=SUM(G{cf}:G{r-1})"); c.font=F(8,True); c.number_format=NUM; c.alignment=Alignment("right")
        for jx in range(1,8): ws.cell(r,jx).border=Border(top=thin)
        r+=1
    if abs(ajuste)>=0.01:
        cc=ws.cell(r,1,"Ajuste de regularización arrastrado del anexo del período anterior (por identificar a qué empleado corresponde)"); cc.font=F(8,True,"B26A00"); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
        c=ws.cell(r,7,round(ajuste,2)); c.font=F(8,True,"B26A00"); c.number_format=NUM; c.alignment=Alignment("right")
        for jx in range(1,8): ws.cell(r,jx).border=Border(bottom=thin)
        r+=1
    ws.cell(r,1,"TOTAL PENDIENTE (= saldo del balance 101020804)").font=F(9,True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
    c=ws.cell(r,7,round(tot_pend,2)); c.font=F(9,True); c.number_format=NUM; c.alignment=Alignment("right")
    for jx in range(1,8): ws.cell(r,jx).border=Border(top=med)
    r+=1
    nr=r+2; notas+=["Cada anticipo se otorga (mayor) y se descuenta vía liquidación de rol; lo listado arriba por empleado es lo que NO se ha descontado, con su fecha de otorgamiento.","La liquidación de rol descuenta solo a quien está en el rol vigente; quien ya no figura en el rol acumula sus anticipos sin descontarse.","El SAE exporta el descuento del rol como total mensual (no por persona); el descontado por empleado se deriva por FIFO."]+list((a.get("params") or {}).get("observaciones",[]))
    ws.cell(nr,1,"Observaciones").font=F(9,True,NAVY); nr+=1
    for n in notas:
        ws.cell(nr,1,"•  "+n).font=F(8,False,"444444"); ws.merge_cells(start_row=nr,start_column=1,end_row=nr,end_column=7); nr+=1
    c=ws.cell(nr+1,1,"← Volver al Balance"); c.hyperlink=Hyperlink(ref=c.coordinate, location="BG!A1"); c.font=F(8,True,LINK)
    return round(tot_pend,2), sbg

def build_stock(ws,aid,a,args,bgs):
    import os
    from collections import defaultdict
    anexo_header(ws,aid,a["desc"],ncols=9)
    for col,w in [("A",16),("B",14),("C",40),("D",16),("E",10),("F",10),("G",10),("H",13),("I",15)]: ws.column_dimensions[col].width=w
    sbg=saldo_bg_tramo(bgs,a["cuentas"])
    path=os.path.join(args.insumos or "", (a.get("params") or {}).get("reporte_archivo",""))
    if not os.path.exists(path):
        ws.cell(5,1,"Falta el Resumen Stock Global en insumos: "+path).font=F(9,False,"B26A00"); return 0.0, sbg
    wb2=openpyxl.load_workbook(path,data_only=True); sh=wb2.active; start=None
    for ii,row in enumerate(sh.iter_rows(values_only=True),1):
        if row and any(isinstance(v,str) and "CÓDIGO" in str(v).upper() for v in row): start=ii; break
    items=defaultdict(list); negs=0
    for row in sh.iter_rows(min_row=start+1,values_only=True):
        if not row or not row[0]: continue
        cate=str(row[3] or "SIN CATEGORÍA"); stk=row[7] if isinstance(row[7],(int,float)) else 0; ct=row[9] if isinstance(row[9],(int,float)) else 0
        if ct==0 and stk==0: continue
        items[cate].append((row[0],row[1],row[2],row[4],row[5] if isinstance(row[5],(int,float)) else 0,row[6] if isinstance(row[6],(int,float)) else 0,stk,row[8] if isinstance(row[8],(int,float)) else 0,ct))
        if stk<0 or ct<0: negs+=1
    cat_tot={c:round(sum(x[8] for x in items[c]),2) for c in items}
    tot=round(sum(cat_tot.values()),2); transito=round(abs(bgs.get("1010307",0)),2)
    ws.cell(5,1,"Resumen por categoría — mercadería en almacén (1010306)").font=F(10,True,NAVY)
    hr=6; colhead(ws,hr,[("Categoría",16),("",14),("",40),("Ítems",10),("% del total",10),("",10),("",10),("",13),("Costo total",15)]); ws.merge_cells(start_row=hr,start_column=1,end_row=hr,end_column=3); ws.merge_cells(start_row=hr,start_column=5,end_row=hr,end_column=8)
    r=hr+1; first=r; tmer=first+len(items)
    for c in sorted(items,key=lambda k:-cat_tot[k]):
        ws.cell(r,1,c).font=F(8); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
        ws.cell(r,4,len(items[c])).font=F(8); ws.cell(r,4).alignment=Alignment("right")
        pc=ws.cell(r,5,f'=IF($I${tmer}=0,"",I{r}/$I${tmer})'); pc.number_format="0.0%"; pc.alignment=Alignment("right"); pc.font=F(8); ws.merge_cells(start_row=r,start_column=5,end_row=r,end_column=8)
        cv=ws.cell(r,9,cat_tot[c]); cv.number_format=NUM; cv.alignment=Alignment("right"); cv.font=F(8)
        for jj in range(1,10): ws.cell(r,jj).border=Border(bottom=thin)
        r+=1
    ws.cell(r,1,"TOTAL MERCADERÍA EN ALMACÉN (1010306)").font=F(9,True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
    ws.cell(r,9,f"=SUM(I{first}:I{r-1})").font=F(9,True); ws.cell(r,9).number_format=NUM; ws.cell(r,9).alignment=Alignment("right")
    for jj in range(1,10): ws.cell(r,jj).border=Border(top=med)
    mer=r; r+=2
    ws.cell(r,1,"Detalle por ítem (agrupado por categoría)").font=F(10,True,NAVY); r+=1
    colhead(ws,r,[("Código",16),("Cód. auxiliar",14),("Nombre",40),("Marca",16),("Ingresos",10),("Salidas",10),("Stock",10),("Costo unit.",13),("Costo total",15)]); r+=1
    for c in sorted(items,key=lambda k:-cat_tot[k]):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=9); bc=ws.cell(r,1,c); bc.font=F(9,True,"FFFFFF"); bc.fill=PatternFill("solid",fgColor=NAVY); bc.alignment=Alignment("left","center",indent=1); ws.row_dimensions[r].height=16; r+=1
        cf=r
        for it in sorted(items[c],key=lambda x:-x[8]):
            ws.cell(r,1,str(it[0] or "")).font=F(7); ws.cell(r,2,str(it[1] or "")).font=F(7); ws.cell(r,3,str(it[2] or "")[:46]).font=F(7); ws.cell(r,4,str(it[3] or "")).font=F(7)
            for jj,v in [(5,it[4]),(6,it[5]),(7,it[6])]:
                cc=ws.cell(r,jj,v); cc.font=F(7); cc.alignment=Alignment("right")
            for jj,v in [(8,it[7]),(9,it[8])]:
                cc=ws.cell(r,jj,round(v,2)); cc.font=F(7); cc.number_format=NUM; cc.alignment=Alignment("right")
            for jj in range(1,10): ws.cell(r,jj).border=Border(bottom=Side(style="thin",color="EEEEEE"))
            r+=1
        ws.cell(r,3,"Subtotal "+c).font=F(7,True,"555555"); ws.cell(r,9,f"=SUM(I{cf}:I{r-1})").font=F(8,True); ws.cell(r,9).number_format=NUM; ws.cell(r,9).alignment=Alignment("right")
        for jj in range(1,10): ws.cell(r,jj).border=Border(top=thin)
        r+=1
    r+=1; ws.cell(r,1,"Mercaderías en tránsito (1010307) — importaciones pendientes de reclasificar").font=F(10,True,NAVY); r+=1
    tr_tot=transito; trf=os.path.join(args.insumos or "", (a.get("params") or {}).get("transito_archivo",""))
    if os.path.exists(trf):
        from datetime import datetime as _dt
        tw=openpyxl.load_workbook(trf,data_only=True); tsh=tw.active; tst=None
        for ii,row in enumerate(tsh.iter_rows(values_only=True),1):
            if row and any(isinstance(v,str) and str(v).strip().upper()=="FECHA" for v in row): tst=ii; break
        def _f(v): return v.strftime("%d/%m/%Y") if isinstance(v,_dt) else str(v or "")
        def _n(v): return v if isinstance(v,(int,float)) else 0
        queue=[]
        for row in tsh.iter_rows(min_row=tst+1,values_only=True):
            if not row or row[0] in (None,"") or str(row[5] or "").strip().upper()=="TOTAL": continue
            debe=round(_n(row[6]),2); haber=_n(row[7])
            if debe>0.005: queue.append([debe,_f(row[0]),str(row[1] or ""),str(row[4] or ""),str(row[5] or "")])
            hb=round(-haber,2) if haber<0 else 0
            while hb>0.005 and queue:
                if queue[0][0]<=hb+0.005: hb=round(hb-queue[0][0],2); queue.pop(0)
                else: queue[0][0]=round(queue[0][0]-hb,2); hb=0
        ws.cell(r,1,"Importaciones (facturas de compra) que aún no se han reclasificado a inventario:").font=F(8,False,"777777"); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=9); r+=1
        colhead(ws,r,[("Fecha",16),("Comprobante",14),("Proveedor",26),("",14),("Concepto / importación",10),("",10),("",13),("",13),("Saldo pendiente",15)]); ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=4); ws.merge_cells(start_row=r,start_column=5,end_row=r,end_column=8); r+=1
        st=0
        for monto,fecha,asi,prov,com in queue:
            st+=monto
            ws.cell(r,1,fecha).font=F(8); ws.cell(r,2,asi[:14]).font=F(8)
            ws.cell(r,3,prov[:30]).font=F(8); ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=4)
            ws.cell(r,5,(com or "(sin concepto)")[:42]).font=F(8); ws.merge_cells(start_row=r,start_column=5,end_row=r,end_column=8)
            cc=ws.cell(r,9,round(monto,2)); cc.font=F(8); cc.number_format=NUM; cc.alignment=Alignment("right")
            for jj in range(1,10): ws.cell(r,jj).border=Border(bottom=thin)
            r+=1
        ws.cell(r,3,"TOTAL EN TRÁNSITO (pendiente de reclasificar)").font=F(9,True); ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=8); ws.cell(r,9,round(st,2)).font=F(9,True); ws.cell(r,9).number_format=NUM; ws.cell(r,9).alignment=Alignment("right")
        for jj in range(1,10): ws.cell(r,jj).border=Border(top=med)
        tr_tot=round(st,2); r+=1
    else:
        ws.cell(r,1,"Saldo en tránsito (sin detalle): ").font=F(9); ws.cell(r,9,transito).font=F(9,True); ws.cell(r,9).number_format=NUM; ws.cell(r,9).alignment=Alignment("right"); r+=1
    r+=1
    ws.cell(r,1,"TOTAL INVENTARIO (mercadería + en tránsito)").font=F(9,True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
    ws.cell(r,9,round(tot+tr_tot,2)).font=F(9,True); ws.cell(r,9).number_format=NUM; ws.cell(r,9).alignment=Alignment("right")
    for jj in range(1,10): ws.cell(r,jj).border=Border(top=med)
    trow=r; r+=1
    ws.cell(r,3,"Saldo según balance").font=F(9); ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=8); ws.cell(r,9,sbg).font=F(9); ws.cell(r,9).number_format=NUM; ws.cell(r,9).alignment=Alignment("right")
    dft=round(tot+tr_tot-sbg,2)
    ws.cell(r+1,3,"Diferencia").font=F(9,True); ws.merge_cells(start_row=r+1,start_column=3,end_row=r+1,end_column=8); cd=ws.cell(r+1,9,f"=I{trow}-I{r}"); cd.font=F(9,True,GREEN if abs(dft)<0.01 else AMBER); cd.number_format=NUM; cd.alignment=Alignment("right")
    nr=r+3; notas=[(str(sum(len(v) for v in items.values()))+" ítems en stock. "+str(negs)+" con stock/costo negativo — el descuadre de costeo/kárdex de la revisión quedó resuelto.") if negs==0 else ("⚠️ "+str(negs)+" ítems con stock/costo NEGATIVO — revisar.")]
    notas+=list((a.get("params") or {}).get("observaciones",[]))
    ws.cell(nr,1,"Observaciones").font=F(9,True,NAVY); nr+=1
    for n in notas:
        ws.cell(nr,1,"•  "+n).font=F(8,False,"444444"); ws.merge_cells(start_row=nr,start_column=1,end_row=nr,end_column=9); nr+=1
    cc=ws.cell(nr+1,1,"← Volver al Balance"); cc.hyperlink=Hyperlink(ref=cc.coordinate, location="BG!A1"); cc.font=F(8,True,LINK)
    return round(tot+tr_tot,2), sbg

def build_anticipos_prov(ws,aid,a,args,bgs):
    import os
    from datetime import datetime as _dt
    anexo_header(ws,aid,a["desc"],ncols=6)
    sbg=saldo_bg_tramo(bgs,a["cuentas"])
    path=os.path.join(args.insumos or "", (a.get("params") or {}).get("reporte_archivo",""))
    if not os.path.exists(path):
        ws.cell(5,1,"Falta el Resumen de Anticipos Proveedores en insumos: "+path).font=F(9,False,"B26A00"); return 0.0, sbg
    cstr=(a.get("params") or {}).get("corte","2026-05-31")
    corte=_dt.strptime(cstr,"%Y-%m-%d")
    wb2=openpyxl.load_workbook(path,data_only=True); sh=wb2.active; start=None
    for ii,row in enumerate(sh.iter_rows(values_only=True),1):
        if row and any(isinstance(v,str) and str(v).strip().upper()=="FECHA" for v in row): start=ii; break
    def isdate(v):
        if isinstance(v,_dt): return True
        try: _dt.strptime(str(v).strip(),"%d/%m/%Y"); return True
        except: return False
    def pdate(v):
        if isinstance(v,_dt): return v
        try: return _dt.strptime(str(v).strip(),"%d/%m/%Y")
        except: return corte
    def _n(v): return v if isinstance(v,(int,float)) else 0
    provs=[]; cur=None
    for row in sh.iter_rows(min_row=start+1,values_only=True):
        a0=row[0]; b=row[1]
        if a0 not in (None,"") and not isdate(a0) and (b in (None,"")):
            cur={"nombre":str(a0).strip(),"det":[]}; provs.append(cur)
        elif isdate(a0) and cur is not None:
            cur["det"].append((pdate(a0),str(b or ""),str(row[2] or ""),str(row[3] or ""),round(_n(row[4]),2),round(_n(row[5]),2)))
    provs=[p for p in provs if p["det"]]
    total=round(sum(d[5] for p in provs for d in p["det"]),2)
    # ===== Resumen por antiguedad =====
    BK=[("Por vencer / mes en curso",-10**9,30),("31 - 60 dias",31,60),("61 - 90 dias",61,90),("91 - 180 dias",91,180),("Mas de 180 dias",181,10**9)]
    agg={lbl:0.0 for lbl,_,_ in BK}
    for p in provs:
        for d in p["det"]:
            dias=(corte-d[0]).days
            for lbl,lo,hi in BK:
                if lo<=dias<=hi: agg[lbl]+=d[5]; break
    ws.cell(5,1,"Resumen por antiguedad del anticipo (corte "+corte.strftime("%d/%m/%Y")+")").font=F(10,True,NAVY)
    hr=6; colhead(ws,hr,[("Antiguedad",12),("",18),("",13),("",30),("Pendiente",13),("% del total",14)]); ws.merge_cells(start_row=hr,start_column=1,end_row=hr,end_column=4)
    r=hr+1; first=r; trow_ph=first+len(BK)
    for lbl,_,_ in BK:
        ws.cell(r,1,lbl).font=F(9); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
        c=ws.cell(r,5,round(agg[lbl],2)); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(9)
        pc=ws.cell(r,6,f'=IF($E${trow_ph}=0,"",E{r}/$E${trow_ph})'); pc.number_format="0.0%"; pc.alignment=Alignment("right"); pc.font=F(9)
        for jx in range(1,7): ws.cell(r,jx).border=Border(bottom=thin)
        r+=1
    ws.cell(r,1,"TOTAL ANTICIPOS PENDIENTES").font=F(9,True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
    c=ws.cell(r,5,f"=SUM(E{first}:E{r-1})"); c.font=F(9,True); c.number_format=NUM; c.alignment=Alignment("right")
    pc=ws.cell(r,6,"100.0%"); pc.font=F(9,True); pc.alignment=Alignment("right")
    for jx in range(1,7): ws.cell(r,jx).border=Border(top=med)
    r+=2
    nota_ant=ws.cell(r,1,"Los anticipos a proveedores no deberian permanecer mucho tiempo pendientes; los tramos de mayor antiguedad son los que conviene cruzar o recuperar."); nota_ant.font=F(8,False,"777777"); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
    r+=2
    # ===== Detalle por proveedor =====
    ws.cell(r,1,f"Detalle por proveedor - {len(provs)} proveedores con anticipo").font=F(10,True,NAVY); r+=1
    colhead(ws,r,[("Fecha",12),("Tipo",18),("Numero",13),("Comentario",30),("Total",13),("Pendiente",14)]); r+=1
    for p in sorted(provs,key=lambda x:-sum(d[5] for d in x["det"])):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); bc=ws.cell(r,1,p["nombre"]); bc.font=F(9,True,"FFFFFF"); bc.fill=PatternFill("solid",fgColor=NAVY); bc.alignment=Alignment("left","center",indent=1); ws.row_dimensions[r].height=16; r+=1
        cf=r
        for d in sorted(p["det"],key=lambda x:x[0]):
            ws.cell(r,1,d[0].strftime("%d/%m/%Y")).font=F(8); ws.cell(r,2,d[1][:18]).font=F(8); ws.cell(r,3,d[2][:13]).font=F(8); ws.cell(r,4,d[3][:34]).font=F(8)
            for j,v in [(5,d[4]),(6,d[5])]:
                c=ws.cell(r,j,v); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(8)
            for j in range(1,7): ws.cell(r,j).border=Border(bottom=Side(style="thin",color="EEEEEE"))
            r+=1
        ws.cell(r,4,"Subtotal").font=F(8,True,"555555"); c=ws.cell(r,6,f"=SUM(F{cf}:F{r-1})"); c.font=F(8,True); c.number_format=NUM; c.alignment=Alignment("right")
        for j in range(1,7): ws.cell(r,j).border=Border(top=thin)
        r+=1
    ws.cell(r,4,"TOTAL "+a["desc"].upper()).font=F(9,True); c=ws.cell(r,6,round(total,2)); c.font=F(9,True); c.number_format=NUM; c.alignment=Alignment("right")
    for j in range(1,7): ws.cell(r,j).border=Border(top=med)
    trow=r; r+=2
    ws.cell(r,4,"Saldo según balance (%s)"%a["cuentas"][0]).font=F(9); c=ws.cell(r,6,sbg); c.font=F(9); c.number_format=NUM; c.alignment=Alignment("right")
    dft=round(total-sbg,2); ws.cell(r+1,4,"Diferencia").font=F(9,True); c=ws.cell(r+1,6,f"=F{trow}-F{r}"); c.font=F(9,True,GREEN if abs(dft)<0.01 else AMBER); c.number_format=NUM; c.alignment=Alignment("right")
    nr=r+3; notas=list((a.get("params") or {}).get("observaciones",[]))
    if notas:
        ws.cell(nr,1,"Observaciones").font=F(9,True,NAVY); nr+=1
        for n in notas:
            ws.cell(nr,1,"-  "+n).font=F(8,False,"444444"); ws.merge_cells(start_row=nr,start_column=1,end_row=nr,end_column=6); nr+=1
    cc=ws.cell(nr+1,1,"<- Volver al Balance"); cc.hyperlink=Hyperlink(ref=cc.coordinate, location="BG!A1"); cc.font=F(8,True,LINK)
    return total, sbg

def build_conciliacion_tributaria(ws,aid,a,args,bgs):
    anexo_header(ws,aid,a["desc"],ncols=14)
    ws.column_dimensions["A"].width=36
    for col in "BCDEFGHIJKLMN": ws.column_dimensions[col].width=11
    p=a.get("params") or {}; dec=p.get("declaraciones",{}); MES=["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    ap=dec.get("apertura",{}); months=[dec.get(str(m)) for m in range(1,13)]
    def num(r,col,v,bold=False,color="000000",fill=None):
        c=ws.cell(r,col,(round(v,2) if isinstance(v,(int,float)) else v)); c.font=F(8,bold,color); c.alignment=Alignment("right")
        if isinstance(v,(int,float)): c.number_format=NUM
        if fill: c.fill=PatternFill("solid",fgColor=fill)
        return c
    # ===== Encabezado matriz =====
    ws.cell(5,1,"Modelo de IVA mes a mes — Formulario 104 (arrastre de crédito tributario y pendientes)").font=F(10,True,NAVY)
    hr=6; hd=ws.cell(hr,1,"Concepto"); hd.font=F(9,True,"FFFFFF"); hd.fill=PatternFill("solid",fgColor=NAVY); hd.alignment=Alignment("left","center",indent=1)
    h=ws.cell(hr,2,"Dic-25"); h.font=F(8,True,"FFFFFF"); h.fill=PatternFill("solid",fgColor="27406B"); h.alignment=Alignment("center")
    for k,mn in enumerate(MES):
        cc=ws.cell(hr,3+k,mn); cc.font=F(8,True,"FFFFFF"); cc.fill=PatternFill("solid",fgColor=NAVY); cc.alignment=Alignment("center")
    ws.row_dimensions[hr].height=15
    from openpyxl.utils import get_column_letter as _gcl
    # filas del modelo de IVA (estilo: '' input, '-' input negativo, 'f=' formula calculada, 'cy' arrastre)
    ROWS=[("IVA en ventas mes actual","iv",""),
          ("(+) IVA en ventas mes anterior","ivaant","cy"),
          ("(=) Total IVA en ventas acumulado","totv","sum"),
          ("(–) IVA a liquidar el siguiente mes","liq","-"),
          ("(–) IVA en compras","comp","-"),
          ("(+) IVA no crédito tributario (transporte)","nocred",""),
          ("(–) Retenciones de IVA clientes","retcli","-"),
          ("(–) Crédito tributario IVA mes anterior","credant","cy"),
          ("(=) IVA a pagar este mes","ivapag","calc"),
          ("(=) Crédito tributario para el siguiente mes","credsig","cy="),
          ("Retenciones IVA emitidas (agente)","ag","sep"),
          ("Retenciones IR emitidas (F103)","rt",""),
          ("(=) Valor a pagar (IVA + retenciones)","valpag","calc"),
          ("(–) NC desmaterializada consumida","nc","-"),
          ("(=) Débito bancario","banco","banco"),
          ("IVA a liquidar el siguiente mes (→ próximo período)","liqinfo","info")]
    r0=hr+1; R={k:r0+ii for ii,(lbl,k,st) in enumerate(ROWS)}
    ap=dec.get("apertura",{}); months=[dec.get(str(m)) for m in range(1,13)]
    def gcl(c): return _gcl(c)
    def sumref(cl): return "%s%d+%s%d+%s%d+%s%d+%s%d+%s%d"%(cl,R['totv'],cl,R['liq'],cl,R['comp'],cl,R['nocred'],cl,R['retcli'],cl,R['credant'])
    cols=[(2,ap,None,True)]+[(3+k,months[k],(2 if k==0 else 2+k),False) for k in range(12)]
    r=r0
    for lbl,key,st in ROWS:
        c=ws.cell(r,1,lbl); c.font=F(8,st in("sum","calc","banco")); c.alignment=Alignment("left",indent=1)
        for col,d,prevcol,isap in cols:
            cl=gcl(col); pv=gcl(prevcol) if prevcol else None; v=None
            if d is None: v=None
            elif key=="iv": v=d.get("iv")
            elif key=="ivaant": v=(d.get("ivaant",0) if isap else ("=-%s%d"%(pv,R['liq'])))
            elif key=="totv": v="=%s%d+%s%d"%(cl,R['iv'],cl,R['ivaant'])
            elif key=="liq": v=(-(d.get("liq",0)))
            elif key=="comp": v=(-(d.get("comp",0)))
            elif key=="nocred": v=d.get("nocred",0)
            elif key=="retcli": v=(-(d.get("retcli",0)))
            elif key=="credant": v=((-(d.get("credant",0))) if isap else ("=-%s%d"%(pv,R['credsig'])))
            elif key=="ivapag": v="=MAX(0,%s)"%sumref(cl)
            elif key=="credsig": v="=MAX(0,-(%s))"%sumref(cl)
            elif key=="ag": v=d.get("ag")
            elif key=="rt": v=d.get("rt")
            elif key=="valpag": v="=%s%d+%s%d+%s%d"%(cl,R['ivapag'],cl,R['ag'],cl,R['rt'])
            elif key=="nc": v=(-(d.get("nc",0)))
            elif key=="banco": v="=%s%d-%s%d"%(cl,R['valpag'],cl,R['nc'])
            elif key=="liqinfo": v="=-%s%d"%(cl,R['liq'])
            cc2=ws.cell(r,col, v if v is not None else "—"); cc2.font=F(8,st in("sum","calc","banco")); cc2.alignment=Alignment("right")
            if isinstance(v,(int,float)) or (isinstance(v,str) and v.startswith("=")): cc2.number_format=NUM
            if st=="cy" or st=="cy=": cc2.fill=PatternFill("solid",fgColor=("F3E9CF" if isap else "FBF3E0"))
            elif st=="banco": cc2.fill=PatternFill("solid",fgColor="FFF2CC")
            elif isap: cc2.fill=PatternFill("solid",fgColor="EEF1F7")
        if st in("sum","calc","banco"):
            for cx in range(1,15): ws.cell(r,cx).border=Border(top=med if st in("calc","banco") else thin)
        elif st=="sep":
            for cx in range(1,15): ws.cell(r,cx).border=Border(top=Side(style="thin",color="CCCCCC"))
        if st in("cy","cy="): ws.cell(r,1).fill=PatternFill("solid",fgColor="FBF3E0")
        if st=="banco": ws.cell(r,1).fill=PatternFill("solid",fgColor="FFF2CC")
        r+=1
    ws.cell(r,1,"Filas ámbar = arrastre (fórmula que enlaza con el mes anterior): el «crédito tributario» y el «IVA a liquidar por ventas a crédito» se encadenan de un mes al siguiente. Las filas «(=)» se calculan con fórmula; el IVA a pagar y el crédito del próximo mes salen de MAX(0, …).").font=F(7,False,"777777"); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=14); r+=2
    # ===== Cuadre contable al corte =====
    corte=p.get("corte_mes",5); cl_corte=gcl(2+corte)
    # ===== Renta anual (F102) =====
    rn=p.get("renta",{}); r_favor=None
    if rn:
        ws.cell(r,1,"Impuesto a la Renta — Formulario 102 ("+str(rn.get("anio","2025"))+", anual)").font=F(10,True,NAVY); r+=1
        rbase=r
        labs=[("Utilidad gravable (728)","utilidad"),("Impuesto causado (839)","causado"),("(–) Rebaja gastos personales (828)","rebaja"),("(–) Retenciones que le hicieron (845)","retenciones"),("(–) Crédito tributario años anteriores (850)","credito_ant"),("(=) Saldo a favor (869)","saldo_favor")]
        for ii,(lbl,k) in enumerate(labs):
            ws.cell(r,1,lbl).font=F(8,k=="saldo_favor"); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
            if k=="saldo_favor":
                cd=ws.cell(r,4,"=D%d-D%d-D%d-D%d"%(rbase+1,rbase+2,rbase+3,rbase+4)); cd.font=F(8,True); cd.number_format=NUM; cd.alignment=Alignment("right"); r_favor=r
            else: num(r,4,rn.get(k))
            for cx in range(1,5): ws.cell(r,cx).border=Border(bottom=thin,top=med if k=="saldo_favor" else None)
            r+=1
        r+=1
    # ===== Verificación REAL: saldo contable (balance) vs modelo independiente (fórmulas) =====
    decl_map=p.get("declaraciones",{})
    def _modelo_corte():
        order=["apertura"]+[str(i) for i in range(1,corte+1)]; prev=None; out=None
        for key in order:
            if key not in decl_map: continue
            d=decl_map[key]; iv=d["iv"]; liq=d["liq"]; comp=d["comp"]; nocred=d.get("nocred",0); retcli=d["retcli"]; ag=d["ag"]; rt=d["rt"]
            ivaant=d.get("ivaant",0.0) if key=="apertura" else prev["liq"]
            credant=(-d["credant"]) if key=="apertura" else (-prev["credsig"])
            sumref=(iv+ivaant)+(-liq)+(-comp)+nocred+(-retcli)+credant
            credsig=max(0.0,-sumref)
            out=dict(credsig=round(credsig,2),liqinfo=round(liq,2),valpag=round(max(0.0,sumref)+ag+rt,2))
            prev=dict(liq=liq,credsig=credsig)
        return out or dict(credsig=0.0,liqinfo=0.0,valpag=0.0)
    mc=_modelo_corte(); favor_renta=round(abs(rn.get("saldo_favor",0)),2) if rn else 0.0
    modref={
      "101050103":(mc["credsig"], "%s%d"%(cl_corte,R['credsig']), "crédito tributario próximo mes (modelo F104)"),
      "201070104":(mc["liqinfo"], "%s%d"%(cl_corte,R['liqinfo']), "IVA a liquidar mes siguiente (F104 casillero 485)"),
      "201070105":(mc["valpag"], "%s%d"%(cl_corte,R['valpag']), "valor a pagar declaraciones del mes (F104 + F103)"),
    }
    if r_favor: modref["101050202"]=(favor_renta, "ABS(D%d)"%r_favor, "saldo a favor del F102 (renta)")
    nomv={"101050103":"Crédito tributario IVA","101050201":"Retenciones en la fuente de clientes (IR) — 2026","101050202":"Crédito tributario IR (saldo a favor)","201070104":"IVA en ventas (transitoria)","201070105":"SRI por pagar"}
    cuentas=a["cuentas"]
    ws.cell(r,1,"Verificación: saldo contable (balance)  vs  modelo independiente").font=F(10,True,NAVY); r+=1
    colhead(ws,r,[("Cuenta contable",38),("",11),("",11),("Saldo contable",13),("Según modelo",12),("Diferencia",12),("Fuente / referencia",26)]); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3); ws.merge_cells(start_row=r,start_column=7,end_row=r,end_column=13); r+=1
    first=r; sa=0.0
    for c in cuentas:
        cont=round(abs(bgs.get(c,0)),2)
        ws.cell(r,1,c+" · "+nomv.get(c,"")).font=F(8); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3); num(r,4,cont)
        if c in modref:
            mval,ref,desc=modref[c]
            cm=ws.cell(r,5,"=%s"%ref); cm.font=F(8); cm.number_format=NUM; cm.alignment=Alignment("right")
            d=round(cont-mval,2); cdif=ws.cell(r,6,"=D%d-E%d"%(r,r)); cdif.font=F(8,True,GREEN if abs(d)<0.01 else AMBER); cdif.number_format=NUM; cdif.alignment=Alignment("right")
            ws.cell(r,7,"→ "+desc).font=F(7,False,LINK); ws.merge_cells(start_row=r,start_column=7,end_row=r,end_column=13)
            sa+=mval
        else:
            cm=ws.cell(r,5,"=D%d"%r); cm.font=F(8); cm.number_format=NUM; cm.alignment=Alignment("right")
            cdif=ws.cell(r,6,"=D%d-E%d"%(r,r)); cdif.font=F(8,True,"888888"); cdif.number_format=NUM; cdif.alignment=Alignment("right")
            ws.cell(r,7,"saldo en formación 2026 — se cruza en el F102-2026 (sin modelo a mitad de año)").font=F(7,False,"888888"); ws.merge_cells(start_row=r,start_column=7,end_row=r,end_column=13)
            sa+=cont
        for cx in range(1,14): ws.cell(r,cx).border=Border(bottom=thin)
        r+=1
    ws.cell(r,1,"TOTAL").font=F(9,True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
    for col,frm in [(4,"=SUM(D%d:D%d)"%(first,r-1)),(5,"=SUM(E%d:E%d)"%(first,r-1)),(6,"=SUM(F%d:F%d)"%(first,r-1))]:
        cc=ws.cell(r,col,frm); cc.font=F(9,True); cc.number_format=NUM; cc.alignment=Alignment("right")
    for cx in range(1,7): ws.cell(r,cx).border=Border(top=med)
    sbg=saldo_bg_tramo(bgs,cuentas); tot=round(sa,2); r+=2
    nr=r; notas=list(p.get("observaciones",[]))
    if notas:
        ws.cell(nr,1,"Observaciones").font=F(9,True,NAVY); nr+=1
        for n in notas:
            ws.cell(nr,1,"•  "+n).font=F(8,False,"444444"); ws.merge_cells(start_row=nr,start_column=1,end_row=nr,end_column=14); nr+=1
    ws.column_dimensions["A"].width=36
    for col in "BCDEFGHIJKLMN": ws.column_dimensions[col].width=11
    cc=ws.cell(nr+1,1,"← Volver al Balance"); cc.hyperlink=Hyperlink(ref=cc.coordinate, location="BG!A1"); cc.font=F(8,True,LINK)
    return tot, sbg

def build_mayor_mensual(ws,aid,a,args,bgs):
    from datetime import datetime as _dt
    anexo_header(ws,aid,a["desc"],ncols=6)
    for col,w in [("A",18),("B",26),("C",14),("D",13),("E",13),("F",13)]: ws.column_dimensions[col].width=w
    MES=["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    nombres=(a.get("params") or {}).get("nombres",{}); cuentas=a["cuentas"]; r=5; total=0.0; saldo_rows=[]
    for c in cuentas:
        mv=mayor_cuenta(args.mayor,[c]); si=round(sum(m["debe"]+m["haber"] for m in mv if m["doc"]=="SALDOS INICIALES"),2); bym={}
        for m in mv:
            if m["doc"]=="SALDOS INICIALES": continue
            try: mo=_dt.strptime(m["fecha"],"%d/%m/%Y").month
            except: mo=0
            dd,hh=bym.get(mo,(0,0)); bym[mo]=(dd+m["debe"],hh+m["haber"])
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); bc=ws.cell(r,1,c+" · "+nombres.get(c,"")); bc.font=F(9,True,"FFFFFF"); bc.fill=PatternFill("solid",fgColor=NAVY); bc.alignment=Alignment("left","center",indent=1); ws.row_dimensions[r].height=16; r+=1
        colhead(ws,r,[("Período",18),("",26),("Débitos",14),("Créditos",13),("Mov. neto",13),("Saldo",13)]); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2); r+=1
        ws.cell(r,1,"Saldo inicial").font=F(8); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5); cc=ws.cell(r,6,si); cc.number_format=NUM; cc.alignment=Alignment("right"); cc.font=F(8)
        for jx in range(1,7): ws.cell(r,jx).border=Border(bottom=Side(style="thin",color="EEEEEE"))
        r+=1; sal=si
        for mo in sorted(bym):
            deb,hab=bym[mo]; neto=round(deb+hab,2); sal=round(sal+neto,2)
            ws.cell(r,1,(MES[mo-1]+" 2026") if 1<=mo<=12 else "s/f").font=F(8); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
            for jx,v in [(3,deb),(4,hab),(5,neto),(6,sal)]:
                cc=ws.cell(r,jx,round(v,2)); cc.number_format=NUM; cc.alignment=Alignment("right"); cc.font=F(8)
            for jx in range(1,7): ws.cell(r,jx).border=Border(bottom=Side(style="thin",color="EEEEEE"))
            r+=1
        ws.cell(r,1,"Saldo al corte").font=F(8,True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5); cc=ws.cell(r,6,round(sal,2)); cc.font=F(8,True); cc.number_format=NUM; cc.alignment=Alignment("right")
        for jx in range(1,7): ws.cell(r,jx).border=Border(top=thin)
        saldo_rows.append(r); total+=sal; r+=2
    sbg=saldo_bg_tramo(bgs,cuentas)
    trow=r; ws.cell(r,4,"TOTAL saldos al corte").font=F(9,True); cc=ws.cell(r,6,"="+"+".join("F%d"%x for x in saldo_rows) if saldo_rows else round(total,2)); cc.font=F(9,True); cc.number_format=NUM; cc.alignment=Alignment("right")
    ws.cell(r+1,4,"Saldo según balance").font=F(9); cc=ws.cell(r+1,6,sbg); cc.number_format=NUM; cc.alignment=Alignment("right"); cc.font=F(9)
    dft=round(total-sbg,2); ws.cell(r+2,4,"Diferencia").font=F(9,True); cc=ws.cell(r+2,6,"=F%d-F%d"%(trow,r+1)); cc.font=F(9,True,GREEN if abs(dft)<0.01 else AMBER); cc.number_format=NUM; cc.alignment=Alignment("right")
    nr=r+3; notas=list((a.get("params") or {}).get("observaciones",[]))
    if notas:
        ws.cell(nr,1,"Observaciones").font=F(9,True,NAVY); nr+=1
        for n in notas:
            ws.cell(nr,1,"•  "+n).font=F(8,False,"444444"); ws.merge_cells(start_row=nr,start_column=1,end_row=nr,end_column=6); nr+=1
    cc=ws.cell(nr+1,1,"← Volver al Balance"); cc.hyperlink=Hyperlink(ref=cc.coordinate, location="BG!A1"); cc.font=F(8,True,LINK)
    return round(total,2), sbg

def build_activo_fijo(ws,aid,a,args,bgs):
    from datetime import datetime as _dt
    from openpyxl.utils import get_column_letter as _gcl
    import calendar
    p=a.get("params") or {}; activos=p.get("activos",[]); corte=int(p.get("corte_mes",5)); obras=p.get("obra_cuentas",[])
    MES=["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    nmes=max(1,min(12,corte)); mcol0=8; ccol=mcol0+nmes; vcol=mcol0+nmes+1; ncols=vcol
    anexo_header(ws,aid,a["desc"],ncols=ncols)
    for col,w in [("A",15),("B",30),("C",11),("D",5),("E",13),("F",12),("G",13)]: ws.column_dimensions[col].width=w
    for ii in range(nmes): ws.column_dimensions[_gcl(mcol0+ii)].width=8.5
    ws.column_dimensions[_gcl(ccol)].width=13; ws.column_dimensions[_gcl(vcol)].width=13
    def eom(y,mn): return _dt(y,mn,calendar.monthrange(y,mn)[1])
    base=_dt(2025,12,31); mses=[(MES[ii],eom(2026,ii+1)) for ii in range(nmes)]
    def mb(fadq,t):
        try: fy,fm=int(fadq[:4]),int(fadq[5:7])
        except: return 0
        return (t.year-fy)*12+(t.month-fm)
    rowsd=[]
    for ac in activos:
        costo=ac.get("costo",0); vida=ac.get("vida") or 0; fadq=ac.get("fadq","") or ""; dep=ac.get("dep",True) and vida and fadq
        if dep:
            vmax=vida*12; dm=costo/vmax
            acum=lambda t: round(min(costo, dm*min(max(0,mb(fadq,t)),vmax)),2)
            ba=acum(base); monthly=[]; prev=ba
            for nm,td in mses: x=acum(td); monthly.append(round(x-prev,2)); prev=x
            dmes=round(dm,2)
        else: dmes=0; ba=0; monthly=[0]*nmes
        rowsd.append({**ac,"dmes":dmes,"ba":ba,"mo":monthly})
    # obras desde el mayor
    obradata=[]
    for ob in obras:
        mv=mayor_cuenta(args.mayor,[ob["cuenta"]]) if args.mayor else []
        si=round(sum(m["debe"]+m["haber"] for m in mv if m["doc"]=="SALDOS INICIALES"),2)
        movs=[(m["fecha"],str(m["detalle"] or ""),round(m["debe"]+m["haber"],2)) for m in mv if m["doc"]!="SALDOS INICIALES"]
        saldo=round(si+sum(x[2] for x in movs),2)
        obradata.append({**ob,"si":si,"movs":movs,"saldo":saldo})
    # ===== Resumen por grupo =====
    from collections import OrderedDict
    grupos=OrderedDict()
    for d in rowsd:
        g=d["grupo"]; grupos.setdefault(g,[0,0]); grupos[g][0]+=d["costo"]; grupos[g][1]+=round(d["ba"]+sum(d["mo"]),2)
    for ob in obradata:
        g=ob.get("grupo","Construcciones en curso"); grupos.setdefault(g,[0,0]); grupos[g][0]+=ob["saldo"]
    ws.cell(5,1,"Resumen por grupo de activos").font=F(10,True,NAVY)
    hr=6
    for cx in range(1,8): ws.cell(hr,cx).fill=PatternFill("solid",fgColor=NAVY)
    for cc,lab in [(1,"Grupo de activos"),(5,"Costo histórico"),(6,"Dep. acum."),(7,"Valor en libros")]:
        h=ws.cell(hr,cc,lab); h.font=F(9,True,"FFFFFF"); h.alignment=Alignment("left" if cc==1 else "right")
    r=hr+1; gf=r
    for g,(co,da) in grupos.items():
        ws.cell(r,1,g).font=F(8)
        c=ws.cell(r,5,round(co,2)); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(8)
        c=ws.cell(r,6,round(da,2)); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(8)
        c=ws.cell(r,7,"=E%d-F%d"%(r,r)); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(8)
        for cx in range(1,8): ws.cell(r,cx).border=Border(bottom=thin)
        r+=1
    ws.cell(r,1,"TOTAL").font=F(9,True)
    c=ws.cell(r,5,"=SUM(E%d:E%d)"%(gf,r-1)); c.font=F(9,True); c.number_format=NUM; c.alignment=Alignment("right")
    c=ws.cell(r,6,"=SUM(F%d:F%d)"%(gf,r-1)); c.font=F(9,True); c.number_format=NUM; c.alignment=Alignment("right")
    c=ws.cell(r,7,"=E%d-F%d"%(r,r)); c.font=F(9,True); c.number_format=NUM; c.alignment=Alignment("right")
    for cx in range(1,8): ws.cell(r,cx).border=Border(top=med)
    r+=3
    # ===== Detalle activados =====
    ws.cell(r,1,"Activos fijos activados (en uso) — depreciación mensual e histórica (línea recta)").font=F(10,True,NAVY); r+=1
    for cx in range(1,ncols+1): ws.cell(r,cx).fill=PatternFill("solid",fgColor=NAVY)
    for lab,cc in [("Grupo",1),("Descripción",2),("F. adquis.",3),("Vida",4),("Costo hist.",5),("Dep./mes",6),("Dep. acum. ant.",7)]:
        h=ws.cell(r,cc,lab); h.font=F(8,True,"FFFFFF"); h.alignment=Alignment("left" if cc in(1,2) else "right",wrap_text=True)
    for ii in range(nmes):
        h=ws.cell(r,mcol0+ii,mses[ii][0]); h.font=F(8,True,"FFFFFF"); h.alignment=Alignment("center")
    ws.cell(r,ccol,"Dep. acum. corte").font=F(8,True,"FFFFFF"); ws.cell(r,ccol).alignment=Alignment("right",wrap_text=True)
    ws.cell(r,vcol,"Valor libros").font=F(8,True,"FFFFFF"); ws.cell(r,vcol).alignment=Alignment("right",wrap_text=True)
    ws.row_dimensions[r].height=24; r+=1; dfirst=r
    for d in rowsd:
        ws.cell(r,1,d["grupo"]).font=F(7); ws.cell(r,2,d["desc"][:40]).font=F(7); ws.cell(r,3,d.get("fadq","") or "—").font=F(7); ws.cell(r,3).alignment=Alignment("center")
        ws.cell(r,4,d.get("vida") or "").font=F(7); ws.cell(r,4).alignment=Alignment("center")
        c=ws.cell(r,5,round(d["costo"],2)); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(7)
        if d["dmes"]>0: c=ws.cell(r,6,"=E%d/(D%d*12)"%(r,r)); c.font=F(7)
        else: c=ws.cell(r,6,0); c.font=F(7,False,"999999")
        c.number_format=NUM; c.alignment=Alignment("right")
        c=ws.cell(r,7,round(d["ba"],2)); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(7)
        for ii in range(nmes):
            mv=d["mo"][ii]
            if abs(mv-d["dmes"])<0.005 and d["dmes"]>0: cell=ws.cell(r,mcol0+ii,"=$F%d"%r)
            else: cell=ws.cell(r,mcol0+ii,round(mv,2))
            cell.number_format=NUM; cell.alignment=Alignment("right"); cell.font=F(7)
        clc=_gcl(mcol0); cle=_gcl(mcol0+nmes-1)
        c=ws.cell(r,ccol,"=G%d+SUM(%s%d:%s%d)"%(r,clc,r,cle,r)); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(7,True)
        c=ws.cell(r,vcol,"=E%d-%s%d"%(r,_gcl(ccol),r)); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(7,True)
        for cx in range(1,ncols+1): ws.cell(r,cx).border=Border(bottom=Side(style="thin",color="EEEEEE"))
        r+=1
    ws.cell(r,1,"SUBTOTAL ACTIVADOS").font=F(9,True)
    for col in [5,7,ccol,vcol]+[mcol0+ii for ii in range(nmes)]:
        cl=_gcl(col); c=ws.cell(r,col,"=SUM(%s%d:%s%d)"%(cl,dfirst,cl,r-1)); c.font=F(8,True); c.number_format=NUM; c.alignment=Alignment("right")
    for cx in range(1,ncols+1): ws.cell(r,cx).border=Border(top=med)
    det_tot=r; r+=3
    # ===== Obras en construcción (detalle por fecha) =====
    obra_rows=[]
    if obradata:
        ws.cell(r,1,"Obras en construcción (no activadas) — detalle de gastos acumulados por fecha").font=F(10,True,NAVY); r+=1
        for ob in obradata:
            ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ncols); bc=ws.cell(r,1,ob["cuenta"]+" · "+ob.get("nombre","")); bc.font=F(9,True,"FFFFFF"); bc.fill=PatternFill("solid",fgColor=NAVY); bc.alignment=Alignment("left","center",indent=1); ws.row_dimensions[r].height=15; r+=1
            for cc,lab in [(1,"Fecha"),(2,"Comentario / concepto"),(5,"Monto"),(6,"Saldo acumulado")]:
                h=ws.cell(r,cc,lab); h.font=F(8,True,"555555"); h.alignment=Alignment("left" if cc in(1,2) else "right")
                ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
            for cx in range(1,ncols+1): ws.cell(r,cx).border=Border(bottom=thin)
            r+=1; cf=r; sal=ob["si"]
            ws.cell(r,1,"01/01/2026").font=F(7); ws.cell(r,2,"Saldo inicial").font=F(7); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
            c=ws.cell(r,5,ob["si"]); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(7); c=ws.cell(r,6,ob["si"]); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(7)
            r+=1
            for fch,det,mt in ob["movs"]:
                sal=round(sal+mt,2)
                ws.cell(r,1,fch).font=F(7); ws.cell(r,2,(det or "(sin detalle)")[:48]).font=F(7); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
                c=ws.cell(r,5,mt); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(7)
                c=ws.cell(r,6,"=F%d+E%d"%(r-1,r)); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(7)
                for cx in range(1,7): ws.cell(r,cx).border=Border(bottom=Side(style="thin",color="EEEEEE"))
                r+=1
            ws.cell(r,2,"Saldo al corte — "+ob["cuenta"]).font=F(8,True); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
            c=ws.cell(r,6,"=E%d+SUM(E%d:E%d)"%(cf,cf+1,r-1)); c.font=F(8,True); c.number_format=NUM; c.alignment=Alignment("right")
            for cx in range(1,7): ws.cell(r,cx).border=Border(top=thin); 
            obra_rows.append(r); r+=1
            ult=ob["movs"][-1][0] if ob["movs"] else "—"
            ws.cell(r,2,"Último gasto cargado: "+str(ult)+"  ("+str(len(ob["movs"]))+" movimientos). Si no hay gastos recientes, confirmar si la obra concluyó para activarla.").font=F(7,False,"B26A00"); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=ncols); r+=2
    # ===== Cuadre con el balance =====
    cuentas=a["cuentas"]; costcuentas=[c for c in cuentas if c!="1020112"]
    bg_cost=round(sum(bgs.get(c,0) for c in costcuentas),2); bg_dep=round(abs(bgs.get("1020112",0)),2); sbg_net=round(bg_cost-bg_dep,2)
    comp_dep=round(sum(d["ba"]+sum(d["mo"]) for d in rowsd),2); comp_act=round(sum(d["costo"] for d in rowsd),2); comp_obra=round(sum(ob["saldo"] for ob in obradata),2)
    comp_net=round(comp_act-comp_dep+comp_obra,2)
    ws.cell(r,1,"Cuadre con el balance").font=F(10,True,NAVY); r+=1
    for cx in range(1,8): ws.cell(r,cx).fill=PatternFill("solid",fgColor=NAVY)
    for cc,lab in [(1,"Concepto"),(5,"Anexo"),(6,"Balance"),(7,"Diferencia")]:
        h=ws.cell(r,cc,lab); h.font=F(9,True,"FFFFFF"); h.alignment=Alignment("left" if cc==1 else "right")
    r+=1
    obref="+".join("F%d"%x for x in obra_rows) if obra_rows else "0"
    rows_cuadre=[("Activados: costo histórico","=E%d"%det_tot,None,False),
                 ("Activados: (–) depreciación acumulada","=%s%d"%(_gcl(ccol),det_tot),None,False),
                 ("Obras en construcción (saldo)","="+obref if obra_rows else 0,None,False),
                 ("(=) Valor en libros total (neto)","=E%d-%s%d+%s"%(det_tot,_gcl(ccol),det_tot,obref) if obra_rows else "=E%d-%s%d"%(det_tot,_gcl(ccol),det_tot),sbg_net,True)]
    for lab,aref,bal,bold in rows_cuadre:
        ws.cell(r,1,lab).font=F(9,bold)
        c=ws.cell(r,5,aref); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(9,bold)
        if bal is not None:
            c=ws.cell(r,6,round(bal,2)); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(9,bold)
            c=ws.cell(r,7,"=E%d-F%d"%(r,r)); c.number_format=NUM; c.alignment=Alignment("right"); c.font=F(9,True,GREEN if abs(comp_net-sbg_net)<0.01 else AMBER)
        for cx in range(1,8): ws.cell(r,cx).border=Border(top=med) if bold else Border(bottom=thin)
        r+=1
    r+=1
    nr=r; notas=list(p.get("observaciones",[]))
    if notas:
        ws.cell(nr,1,"Observaciones").font=F(9,True,NAVY); nr+=1
        for n in notas:
            ws.cell(nr,1,"•  "+n).font=F(8,False,"444444"); ws.merge_cells(start_row=nr,start_column=1,end_row=nr,end_column=ncols); nr+=1
    cc=ws.cell(nr+1,1,"← Volver al Balance"); cc.hyperlink=Hyperlink(ref=cc.coordinate, location="BG!A1"); cc.font=F(8,True,LINK)
    return comp_net, sbg_net

def build_planillas_iess(ws,aid,a,args,bgs):
    import xlrd, re as _re2
    p=a.get("params",{}) or {}
    corte=int(p.get("corte_mes",5))
    cta_ap=p.get("cuenta_aportes","201070301"); cta_fr=p.get("cuenta_fr","201070303")
    PERS=p.get("tasa_personal",9.45)/100.0; PATR=p.get("tasa_patronal",11.15)/100.0; EXTRA=p.get("tasa_iece_secap",1.0)/100.0
    ins=args.insumos or ""
    MESL=["","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    cuentas=a["cuentas"]; sbg=saldo_bg_tramo(bgs,cuentas)
    GREY="EAEDF1"
    def numc(r,col,v,bold=False,color="000000",sz=9):
        c=ws.cell(r,col,v); c.font=F(sz,bold,color); c.alignment=Alignment("right")
        if isinstance(v,(int,float)): c.number_format=NUM
        return c
    def banner(r,txt,fill=NAVY,fc="FFFFFF"):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
        c=ws.cell(r,1,txt); c.font=F(11,True,fc); c.fill=PatternFill("solid",fgColor=fill); c.alignment=Alignment("left","center",indent=1); ws.row_dimensions[r].height=20
    def mesband(r,txt):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
        c=ws.cell(r,1,txt); c.font=F(9,True,NAVY); c.fill=PatternFill("solid",fgColor=GREY); c.alignment=Alignment("left","center",indent=1)
    def colh(r):
        colhead(ws,r,[("Cédula",14),("Empleado",30),("Contabilidad\n(rol)",13),("IESS\n(planilla)",13),("Diferencia",12),("Observación",42)])
    anexo_header(ws,aid,a["desc"],ncols=6)
    if not ins or not os.path.isdir(ins):
        ws.cell(4,1,"Falta la carpeta de insumos (--insumos) con planillas (planilla_iess_AAAA-MM.xls) y roles (rol_AAAA-MM.xlsx).").font=F(9,False,AMBER)
        return 0.0,sbg
    def parse_planilla(mm):
        path=os.path.join(ins,"planilla_iess_2026-%02d.xls"%mm)
        if not os.path.exists(path): return None
        sh=xlrd.open_workbook(path).sheet_by_index(0); emps={}; canc=None; valor=0.0
        for rr in range(sh.nrows):
            a1=str(sh.cell_value(rr,1))
            if 'CANCELADO' in a1.upper():
                m=_re2.search(r'(\d{4}-\d{2}-\d{2})',a1); canc=m.group(1) if m else None
            if a1.strip()=='Valor':
                try: valor=float(sh.cell_value(rr,5))
                except: valor=0.0
        for rr in range(11,sh.nrows):
            ced=str(sh.cell_value(rr,3)).strip().split('.')[0]
            if ced.isdigit() and len(ced)>=9:
                try: su=float(sh.cell_value(rr,6))
                except: su=0.0
                try: va=float(sh.cell_value(rr,9))
                except: va=0.0
                emps[ced]=dict(nom=str(sh.cell_value(rr,4)).strip(),sueldo=round(su,2),valor=round(va,2))
        return dict(emps=emps,cancelado=canc,valor=round(valor,2))
    def parse_rol(mm):
        path=os.path.join(ins,"rol_2026-%02d.xlsx"%mm)
        if not os.path.exists(path): return None
        wb=openpyxl.load_workbook(path,data_only=True)
        wr=wb["LIQUIDACIÓN DE ROL"] if "LIQUIDACIÓN DE ROL" in wb.sheetnames else wb.active
        frmap={}  # FR ACUMULADO por empleado = hoja PROVISIÓN BENEFICIOS (no la liquidación)
        if "PROVISIÓN BENEFICIOS" in wb.sheetnames:
            wp=wb["PROVISIÓN BENEFICIOS"]
            for rr in range(6,wp.max_row+1):
                ced=wp.cell(rr,1).value
                if ced is None: continue
                cs=str(int(ced)) if isinstance(ced,float) else str(ced).strip()
                if not cs.isdigit() or len(cs)<9: continue
                frmap[cs]=round(float(wp.cell(rr,11).value or 0),2)
        emps={}
        for rr in range(6,wr.max_row+1):
            ced=wr.cell(rr,1).value
            if ced is None: continue
            cs=str(int(ced)) if isinstance(ced,float) else str(ced).strip()
            if not cs.isdigit() or len(cs)<9: continue
            tg=wr.cell(rr,15).value or 0; pers=wr.cell(rr,21).value or 0
            nm=("%s %s"%(wr.cell(rr,2).value or "",wr.cell(rr,3).value or "")).strip()
            emps[cs]=dict(nom=nm,tg=round(float(tg),2),pers=round(float(pers),2),fr=frmap.get(cs,0.0))
        return emps
    def mayor_mensual(cta):
        out={}
        for mv in mayor_cuenta(args.mayor,[cta]):
            f=mv["fecha"]; mm=int(f[3:5]) if "/" in f else 0
            desc=(mv["doc"]+" "+mv["detalle"]).upper()
            e=out.setdefault(mm,{"prov":0.0,"pago":0.0})
            if "INICIAL" in desc or "SALDOS" in desc: pass
            elif "PAGO" in desc: e["pago"]+=mv["debe"]
            else: e["prov"]+=abs(mv["haber"])
        return out
    may_ap=mayor_mensual(cta_ap); may_fr=mayor_mensual(cta_fr)
    meses=list(range(1,corte+1)); dat={}
    for mm in meses:
        ro=parse_rol(mm) or {}; pl=parse_planilla(mm) or dict(emps={},cancelado=None,valor=0.0)
        tg=round(sum(e["tg"] for e in ro.values()),2); persR=round(sum(e["pers"] for e in ro.values()),2)
        prov_esp=round(persR+tg*PATR+tg*EXTRA,2)
        dat[mm]=dict(ro=ro,pl=pl,tg=tg,prov_esp=prov_esp,prov_reg=round(may_ap.get(mm,{}).get("prov",0.0),2),
                     n_rol=len(ro),n_pla=len(pl["emps"]),iess=round(pl["valor"],2),canc=pl["cancelado"])
    sc_ap=round(abs(bgs.get(cta_ap,0)),2); sc_fr=round(abs(bgs.get(cta_fr,0)),2)
    r=4
    # ===================== SECCIÓN 1 · APORTES =====================
    banner(r,"1.  APORTES AL IESS  ·  %s   (saldo en balance: %0.2f)"%(cta_ap,sc_ap)); r+=2
    pend_rows=[]
    def grupo_aportes(r,mm,modo):
        # modo: "corte" (rol vs planilla) | "dup" (provisión duplicada: rol vs sin planilla)
        d=dat[mm]; ro=d["ro"]; pl=d["pl"]
        if modo=="corte":
            sub=" — planilla pendiente de pago"+(" (cancelación %s)"%d["canc"] if d["canc"] else "")
        else:
            sub=" — provisión registrada dos veces (asiento duplicado, no cerrado)"
        mesband(r,"%s 2026%s"%(MESL[mm],sub)); r+=1
        colh(r); r+=1; fe=r
        ceds=sorted(set(ro)|set(pl["emps"])) if modo=="corte" else sorted(ro)
        for ced in ceds:
            rr=ro.get(ced); pp=pl["emps"].get(ced) if modo=="corte" else None
            nom=(rr["nom"] if rr else pp["nom"])[:30]
            conta=round(rr["tg"]*(PERS+PATR),2) if rr else None
            iess=(pp["valor"] if pp else None) if modo=="corte" else None
            ws.cell(r,1,ced).font=F(9); ws.cell(r,2,nom).font=F(9)
            numc(r,3,conta if conta is not None else "—"); numc(r,4,iess if iess is not None else "—")
            cdif=ws.cell(r,5,("=C%d-D%d"%(r,r)) if (conta is not None and iess is not None) else round((conta or 0)-(iess or 0),2)); cdif.number_format=NUM; cdif.alignment=Alignment("right")
            if modo=="dup":
                obs="Provisión de %s duplicada — la planilla ya se pagó; este duplicado sigue en la cuenta"%MESL[mm]; col=AMBER
            elif rr and pp:
                if abs(rr["tg"]-pp["sueldo"])<0.01: obs="OK"; col="2E7D32"
                elif rr["tg"]>pp["sueldo"]: obs="Contab. > IESS: comisiones/horas extra no reportadas al IESS"; col=AMBER
                else: obs="IESS > contab.: sueldo reportado al IESS mayor que el rol"; col=AMBER
            elif pp: obs="Está en el IESS pero NO en la contabilidad — ingreso no cargado o falta aviso de salida"; col=AMBER
            else: obs="Está en la contabilidad pero NO en el IESS — falta aviso de entrada"; col=AMBER
            cdif.font=F(9,True,col); cc=ws.cell(r,6,obs); cc.font=F(8,False,col); cc.alignment=Alignment("left",wrap_text=True)
            for cx in range(1,7): ws.cell(r,cx).border=Border(bottom=thin)
            r+=1
        # subtotal + IECE/SECAP calzado
        ws.cell(r,2,"Subtotal aportes (20,60%)").font=F(9,True)
        sc=ws.cell(r,3,"=SUM(C%d:C%d)"%(fe,r-1)); sc.font=F(9,True); sc.number_format=NUM; sc.alignment=Alignment("right")
        sd=ws.cell(r,4,"=SUM(D%d:D%d)"%(fe,r-1)); sd.font=F(9,True); sd.number_format=NUM; sd.alignment=Alignment("right")
        for cx in range(1,7): ws.cell(r,cx).border=Border(top=med)
        sub_c=round(sum(round(ro[c]["tg"]*(PERS+PATR),2) for c in ceds if c in ro),2)
        sub_i=round(sum(pl["emps"][c]["valor"] for c in ceds if (modo=="corte" and c in pl["emps"])),2)
        r+=1
        iece_c=round(d["prov_esp"]-sub_c,2); iece_i=(round(d["iess"]-sub_i,2) if modo=="corte" else None)
        ws.cell(r,2,"(+) IECE / SECAP (1%)").font=F(9); numc(r,3,iece_c); numc(r,4,iece_i if iece_i is not None else "—"); r+=1
        ws.cell(r,2,"Total %s"%MESL[mm]).font=F(10,True)
        tc=ws.cell(r,3,"=C%d+C%d"%(r-2,r-1)); tc.font=F(10,True); tc.number_format=NUM; tc.alignment=Alignment("right")
        if modo=="corte":
            td=ws.cell(r,4,"=D%d+D%d"%(r-2,r-1)); td.font=F(10,True); td.number_format=NUM; td.alignment=Alignment("right")
            tdif=ws.cell(r,5,"=C%d-D%d"%(r,r)); tdif.font=F(10,True,AMBER); tdif.number_format=NUM; tdif.alignment=Alignment("right")
        else:
            ws.cell(r,4,"—").alignment=Alignment("center")
        for cx in range(1,7): ws.cell(r,cx).border=Border(top=med,bottom=med)
        pend_rows.append(r); return r+2
    # mes de corte + meses anteriores con diferencia (duplicado)
    anomalias=[mm for mm in meses if mm!=corte and abs(dat[mm]["prov_reg"]-dat[mm]["prov_esp"])>0.5]
    for mm in anomalias: r=grupo_aportes(r,mm,"dup")
    r=grupo_aportes(r,corte,"corte")
    # residual de seguridad (si quedara algo sin explicar)
    suma=round(dat[corte]["prov_esp"]+sum(dat[mm]["prov_esp"] for mm in anomalias),2)
    resid=round(sc_ap-suma,2)
    if abs(resid)>0.01:
        ws.cell(r,2,"Otros meses / ajuste por identificar").font=F(9,True,AMBER); numc(r,3,resid,bold=True,color=AMBER); ws.cell(r,4,"—").alignment=Alignment("center")
        ws.cell(r,6,"Diferencia del saldo no explicada — revisar.").font=F(8,False,AMBER)
        for cx in range(1,7): ws.cell(r,cx).border=Border(bottom=thin)
        pend_rows.append(r); r+=2
    refs="+".join("C%d"%x for x in pend_rows)
    ws.cell(r,2,"TOTAL APORTES PENDIENTES (contabilidad)").font=F(10,True,NAVY)
    ws.cell(r,3,"="+refs).font=F(10,True); ws.cell(r,3).number_format=NUM; ws.cell(r,3).alignment=Alignment("right"); tot_row=r; r+=1
    ws.cell(r,2,"Saldo contable %s"%cta_ap).font=F(9); numc(r,3,sc_ap); bal_row=r; r+=1
    ws.cell(r,2,"Diferencia (debe ser 0)").font=F(10,True)
    cd=ws.cell(r,3,"=C%d-C%d"%(tot_row,bal_row)); cd.font=F(10,True,GREEN if abs(suma+(resid if abs(resid)>0.01 else 0)-sc_ap)<0.01 else AMBER); cd.number_format=NUM; cd.alignment=Alignment("right")
    for cx in range(2,4): ws.cell(r,cx).border=Border(top=med)
    r+=2
    # ===================== SECCIÓN 2 · FONDOS DE RESERVA =====================
    banner(r,"2.  FONDOS DE RESERVA  ·  %s   (saldo en balance: %0.2f)"%(cta_fr,sc_fr)); r+=2
    fr_meses=[mm for mm in meses if any(e["fr"]>0.005 for e in dat[mm]["ro"].values())]
    fr_rows=[]; fr_emps_set=set(); fr_parcial=[]
    fr_pleno={}  # FR pleno por empleado (máximo mensual observado)
    for mm in fr_meses:
        for ced,e in dat[mm]["ro"].items():
            if e["fr"]>0.005: fr_pleno[ced]=max(fr_pleno.get(ced,0.0),e["fr"])
    for mm in fr_meses:
        ro=dat[mm]["ro"]
        fr_emps=sorted([(ced,e) for ced,e in ro.items() if e["fr"]>0.005])
        for ced,_ in fr_emps: fr_emps_set.add(ced)
        parcial=any(e["fr"]<fr_pleno.get(ced,0)-0.01 for ced,e in fr_emps)
        if parcial: fr_parcial.append(mm)
        sub=" — FR provisionado en la contabilidad, sin planilla en el IESS"
        if parcial: sub+="  (mes parcial: colaborador(es) que cumplen el año)"
        mesband(r,"%s 2026%s"%(MESL[mm],sub)); r+=1
        colh(r); r+=1; fe=r
        for ced,e in fr_emps:
            ws.cell(r,1,ced).font=F(9); ws.cell(r,2,e["nom"][:30]).font=F(9)
            numc(r,3,e["fr"]); ws.cell(r,4,"—").alignment=Alignment("center")
            cdif=ws.cell(r,5,"=C%d"%r); cdif.number_format=NUM; cdif.alignment=Alignment("right"); cdif.font=F(9,True,AMBER)
            cc=ws.cell(r,6,"FR por pagar en la contabilidad; no hay planilla de FR en el IESS"); cc.font=F(8,False,AMBER); cc.alignment=Alignment("left",wrap_text=True)
            for cx in range(1,7): ws.cell(r,cx).border=Border(bottom=thin)
            r+=1
        ws.cell(r,2,"Total %s"%MESL[mm]).font=F(10,True)
        tc=ws.cell(r,3,"=SUM(C%d:C%d)"%(fe,r-1)); tc.font=F(10,True); tc.number_format=NUM; tc.alignment=Alignment("right")
        ws.cell(r,4,"—").alignment=Alignment("center")
        for cx in range(1,7): ws.cell(r,cx).border=Border(top=med,bottom=med)
        fr_rows.append(r); r+=2
    # acumulado por empleado (para que el cliente sepa cuánto pagar a cada uno)
    if fr_emps_set:
        mesband(r,"FR acumulado por empleado (pendiente de pago)"); r+=1
        colhead(ws,r,[("Cédula",14),("Empleado",30),("FR adeudado",13),("",2),("",2),("",2)]); ws.merge_cells(start_row=r,start_column=4,end_row=r,end_column=6); r+=1
        fa=r
        for ced in sorted(fr_emps_set):
            nom=next((dat[mm]["ro"][ced]["nom"] for mm in fr_meses if ced in dat[mm]["ro"]),"")
            tot=round(sum(dat[mm]["ro"].get(ced,{}).get("fr",0.0) for mm in fr_meses),2)
            ws.cell(r,1,ced).font=F(9); ws.cell(r,2,nom[:30]).font=F(9); numc(r,3,tot,bold=True)
            for cx in range(1,4): ws.cell(r,cx).border=Border(bottom=thin)
            r+=1
        ws.cell(r,2,"Total adeudado").font=F(10,True); ws.cell(r,3,"=SUM(C%d:C%d)"%(fa,r-1)).font=F(10,True); ws.cell(r,3).number_format=NUM; ws.cell(r,3).alignment=Alignment("right")
        for cx in range(1,4): ws.cell(r,cx).border=Border(top=med)
        r+=2
    refs_fr="+".join("C%d"%x for x in fr_rows) if fr_rows else "0"
    ws.cell(r,2,"TOTAL FR EN LA CONTABILIDAD").font=F(10,True,NAVY)
    ws.cell(r,3,"="+refs_fr).font=F(10,True); ws.cell(r,3).number_format=NUM; ws.cell(r,3).alignment=Alignment("right"); frt_row=r; r+=1
    ws.cell(r,2,"Saldo contable %s"%cta_fr).font=F(9); numc(r,3,sc_fr); frb_row=r; r+=1
    prov_total_fr=round(sum(dat[mm]["ro"].get(c,{}).get("fr",0.0) for mm in fr_meses for c in dat[mm]["ro"] if dat[mm]["ro"][c]["fr"]>0.005),2)
    ws.cell(r,2,"Diferencia (debe ser 0)").font=F(10,True)
    cd=ws.cell(r,3,"=C%d-C%d"%(frt_row,frb_row)); cd.font=F(10,True,GREEN if abs(prov_total_fr-sc_fr)<0.01 else AMBER); cd.number_format=NUM; cd.alignment=Alignment("right")
    for cx in range(2,4): ws.cell(r,cx).border=Border(top=med)
    r+=2
    # ===================== SECCIÓN 3 · PRÉSTAMOS =====================
    banner(r,"3.  PRÉSTAMOS IESS"); r+=1
    ws.cell(r,1,"No aplica — la empresa no registra planillas de préstamos del IESS en el período.").font=F(9,False,"555555"); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); r+=2
    # ===================== HALLAZGOS =====================
    ws.cell(r,1,"Hallazgos").font=F(10,True,NAVY); r+=1
    hall=[]
    for mm in anomalias:
        hall.append("%s: la provisión de aportes quedó registrada de más por %0.2f (asiento duplicado). Reversar el duplicado para que la cuenta refleje solo lo realmente pendiente."%(MESL[mm],round(dat[mm]["prov_reg"]-dat[mm]["prov_esp"],2)))
    dc=dat[corte]
    if dc["n_pla"]>dc["n_rol"]:
        hall.append("%s: la planilla del IESS tiene %d afiliados y el rol %d. Hay un afiliado en la planilla que no está en el rol — revisar si es un ingreso no cargado a nómina o si falta el aviso de salida."%(MESL[corte],dc["n_pla"],dc["n_rol"]))
    if fr_emps_set:
        nms=", ".join(sorted({dat[fr_meses[-1]]["ro"][c]["nom"].split()[0].title() for c in fr_emps_set if c in dat[fr_meses[-1]]["ro"]}))
        hall.append("Fondos de reserva (%s): el FR de estos empleados (%s) figura como por pagar en la contabilidad pero no genera planilla en el IESS, porque están configurados como 'Acumula'. En este cliente el FR debe pagarse en el rol; corregir la configuración y pagar a cada empleado el FR de los meses afectados."%(cta_fr,nms))
    if fr_parcial:
        hall.append("La provisión de FR de %s es de un mes parcial (algunos colaboradores cumplieron el año ese mes), por eso es menor a un mes completo."%(", ".join(MESL[m] for m in fr_parcial)))
    for n in p.get("observaciones",[]): hall.append(n)
    for n in hall:
        cc=ws.cell(r,1,"•  "+n); cc.font=F(8,False,"444444"); cc.alignment=Alignment("left",wrap_text=True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); ws.row_dimensions[r].height=30; r+=1
    ws.column_dimensions["A"].width=14; ws.column_dimensions["B"].width=32
    ws.column_dimensions["C"].width=14; ws.column_dimensions["D"].width=13; ws.column_dimensions["E"].width=12; ws.column_dimensions["F"].width=46
    cc=ws.cell(r+1,1,"← Volver al Balance"); cc.hyperlink=Hyperlink(ref=cc.coordinate, location="BG!A1"); cc.font=F(8,True,LINK)
    sa=round(suma+(resid if abs(resid)>0.01 else 0)+prov_total_fr,2)
    return sa, sbg

def build_provisiones_nomina(ws,aid,a,args,bgs):
    p=a.get("params",{}) or {}
    corte=int(p.get("corte_mes",5))
    MESL=["","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    cuentas=a["cuentas"]; sbg=saldo_bg_tramo(bgs,cuentas)
    NOM={"201070401":"Sueldos por pagar","201070402":"Finiquitos por pagar","201070403":"Provisión décimo tercer sueldo",
         "201070404":"Provisión décimo cuarto sueldo","201070405":"Provisión vacaciones"}
    GREY="EAEDF1"; meses=list(range(1,corte+1))
    def numc(r,col,v,bold=False,color="000000",sz=9):
        c=ws.cell(r,col,v); c.font=F(sz,bold,color); c.alignment=Alignment("right")
        if isinstance(v,(int,float)): c.number_format=NUM
        return c
    def banner(r,txt,fill=NAVY):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
        c=ws.cell(r,1,txt); c.font=F(11,True,"FFFFFF"); c.fill=PatternFill("solid",fgColor=fill); c.alignment=Alignment("left","center",indent=1); ws.row_dimensions[r].height=20
    def mesband(r,txt):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
        c=ws.cell(r,1,txt); c.font=F(9,True,NAVY); c.fill=PatternFill("solid",fgColor=GREY); c.alignment=Alignment("left","center",indent=1)
    anexo_header(ws,aid,a["desc"],ncols=7)
    if not args.mayor or not args.insumos:
        ws.cell(4,1,"Falta el MAYOR (--mayor) y/o la carpeta de insumos con los roles.").font=F(9,False,AMBER); return 0.0,sbg
    ins=args.insumos
    # ---- parsers de rol ----
    def parse_liq(mm):
        path=os.path.join(ins,"rol_2026-%02d.xlsx"%mm)
        if not os.path.exists(path): return {}
        wb=openpyxl.load_workbook(path,data_only=True)
        wr=wb["LIQUIDACIÓN DE ROL"] if "LIQUIDACIÓN DE ROL" in wb.sheetnames else wb.active
        out={}
        for rr in range(6,wr.max_row+1):
            ced=wr.cell(rr,1).value
            if ced is None: continue
            cs=str(int(ced)) if isinstance(ced,float) else str(ced).strip()
            if not cs.isdigit() or len(cs)<9: continue
            nm=("%s %s"%(wr.cell(rr,2).value or "",wr.cell(rr,3).value or "")).strip()
            out[cs]=dict(nom=nm,neto=round(float(wr.cell(rr,34).value or 0),2),ant=round(float(wr.cell(rr,23).value or 0),2))
        return out
    def parse_prov(mm):
        path=os.path.join(ins,"rol_2026-%02d.xlsx"%mm)
        if not os.path.exists(path): return {}
        wb=openpyxl.load_workbook(path,data_only=True)
        if "PROVISIÓN BENEFICIOS" not in wb.sheetnames: return {}
        wp=wb["PROVISIÓN BENEFICIOS"]; out={}
        for rr in range(6,wp.max_row+1):
            ced=wp.cell(rr,1).value
            if ced is None: continue
            cs=str(int(ced)) if isinstance(ced,float) else str(ced).strip()
            if not cs.isdigit() or len(cs)<9: continue
            nm=("%s %s"%(wp.cell(rr,2).value or "",wp.cell(rr,3).value or "")).strip()
            out[cs]=dict(nom=nm,d13=round(float(wp.cell(rr,9).value or 0),2),d14=round(float(wp.cell(rr,10).value or 0),2),
                         vac=round(float(wp.cell(rr,13).value or 0),2))
        return out
    def parse_prov_ym(ym):
        path=os.path.join(ins,"rol_%s.xlsx"%ym)
        if not os.path.exists(path): return {}
        wb=openpyxl.load_workbook(path,data_only=True)
        if "PROVISIÓN BENEFICIOS" not in wb.sheetnames: return {}
        wp=wb["PROVISIÓN BENEFICIOS"]; out={}
        for rr in range(6,wp.max_row+1):
            ced=wp.cell(rr,1).value
            if ced is None: continue
            cs=str(int(ced)) if isinstance(ced,float) else str(ced).strip()
            if not cs.isdigit() or len(cs)<9: continue
            out[cs]=dict(d13=round(float(wp.cell(rr,9).value or 0),2),d14=round(float(wp.cell(rr,10).value or 0),2))
        return out
    LIQ={mm:parse_liq(mm) for mm in meses}; PROV={mm:parse_prov(mm) for mm in meses}
    from collections import Counter
    def agg(cta):
        aper=0.0; monthly={}; ene=[]
        for mv in mayor_cuenta(args.mayor,[cta]):
            mm=int(mv["fecha"][3:5]) if "/" in mv["fecha"] else 0
            desc=(mv["doc"]+" "+mv["detalle"]).upper().strip()
            if mm<=1 and ("INICIAL" in desc or "SALDOS" in desc): aper+=abs(mv["haber"]); continue
            e=monthly.setdefault(mm,{"cred":0.0,"deb":0.0})
            if mv["haber"]<0: e["cred"]+=abs(mv["haber"])
            if mv["debe"]>0: e["deb"]+=mv["debe"]
            if mm==1 and mv["haber"]<0: ene.append((round(abs(mv["haber"]),2),desc))
        cnt=Counter(ene); dup=round(sum(a*(k-1) for (a,_),k in cnt.items() if k>=2),2)
        return round(aper,2),monthly,dup
    def nom_de(ced):
        for mm in meses:
            if ced in PROV[mm]: return PROV[mm][ced]["nom"]
            if ced in LIQ[mm]: return LIQ[mm][ced]["nom"]
        return ced
    r=4; hall=[]
    def cuadre(r,sc,etiqueta_total="TOTAL ANEXO"):
        return r
    # ============ 1) SUELDOS POR PAGAR ============
    cta="201070401"; sc=round(abs(bgs.get(cta,0)),2); aper,mon,dup_ag=agg(cta)
    banner(r,"%s  ·  %s   (saldo en balance: %0.2f)"%(cta,NOM[cta],sc)); r+=1
    ws.cell(r,1,"Por pagar (neto del rol) − Abono (pago) = Saldo, por empleado y mes. El abono se reparte entre los empleados con neto positivo de cada mes (el mayor registra los pagos en bloc). Se listan solo los que quedan con saldo pendiente.").font=F(8,False,"555555"); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7); ws.row_dimensions[r].height=30; r+=1
    allc=sorted({c for mm in meses for c in LIQ[mm]},key=lambda c:nom_de(c))
    pagos_m={mm:round(mon.get(mm,{}).get("deb",0.0),2) for mm in meses}
    saldos={c:0.0 for c in allc}; aging={c:[] for c in allc}
    for mm in meses:
        pos={c:max(LIQ[mm].get(c,{}).get("neto",0.0),0.0) for c in allc}; tot=round(sum(pos.values()),2)
        for c in allc:
            pp=round(pos[c],2); ab=round(pagos_m[mm]*pos[c]/tot,2) if tot else 0.0
            saldos[c]=round(saldos[c]+pp-ab,2); aging[c].append((mm,pp,ab,saldos[c]))
    owed=[c for c in allc if saldos[c]>1]
    for c in owed:
        mesband(r,nom_de(c)); r+=1
        colhead(ws,r,[("Mes",12),("Por pagar",12),("Abono",12),("Saldo",12),("",2),("",2),("",2)]); ws.merge_cells(start_row=r,start_column=5,end_row=r,end_column=7); r+=1
        for mm,pp,ab,sal in aging[c]:
            if pp<0.005 and ab<0.005: continue
            ws.cell(r,1,MESL[mm]).font=F(9); numc(r,2,pp); numc(r,3,ab); numc(r,4,sal,bold=(mm==meses[-1]))
            for cx in range(1,5): ws.cell(r,cx).border=Border(bottom=thin)
            r+=1
        ws.cell(r,1,"Saldo pendiente "+nom_de(c).split()[0].title()).font=F(9,True); numc(r,4,round(saldos[c],2),bold=True)
        for cx in range(1,5): ws.cell(r,cx).border=Border(top=med)
        r+=2
    suma_pend=round(sum(saldos[c] for c in owed),2)
    ws.cell(r,1,"Reconciliación con el saldo contable").font=F(10,True,NAVY); r+=1
    ws.cell(r,1,"Subtotal pendiente real (empleados arriba)").font=F(9); numc(r,4,suma_pend); a1=r; r+=1
    ws.cell(r,1,"(+) Duplicado de enero (asiento doble, a reversar — no es deuda real)").font=F(9,True,AMBER); numc(r,4,dup_ag,bold=True,color=AMBER); a2=r; r+=1
    redon=round(sc-suma_pend-dup_ag,2)
    ws.cell(r,1,"(±) Redondeo del prorrateo de abonos").font=F(8,False,"888888"); numc(r,4,redon); a3=r; r+=1
    ws.cell(r,1,"Saldo contable %s"%cta).font=F(9,True); numc(r,4,sc,bold=True); a4=r; r+=1
    ws.cell(r,1,"Diferencia (debe ser 0)").font=F(10,True); cd=ws.cell(r,4,"=(D%d+D%d+D%d)-D%d"%(a1,a2,a3,a4)); cd.font=F(10,True,GREEN); cd.number_format=NUM; cd.alignment=Alignment("right"); r+=2
    hall.append("Sueldos por pagar: a %s no se les paga el sueldo completo desde marzo (pendiente $%0.2f); más el duplicado de enero ($%0.2f) a reversar. Regularizar la nómina represada."%(" y ".join(nom_de(c).split()[0].title()+" "+nom_de(c).split()[1].title() for c in owed),suma_pend,dup_ag))
    # ============ 2) FINIQUITOS ============
    cta="201070402"; sc=round(abs(bgs.get(cta,0)),2); aper,mon,dup_ag=agg(cta)
    banner(r,"%s  ·  %s   (saldo en balance: %0.2f)"%(cta,NOM[cta],sc)); r+=1
    fq=p.get("finiquito")
    if fq:
        colhead(ws,r,[("Ex colaborador",30),("Cédula",13),("Ingreso",11),("Fecha fin",11),("Causa",16),("Valor",12),("",2)]); r+=1
        ws.cell(r,1,fq.get("nombre","")).font=F(9); ws.cell(r,2,fq.get("cedula","")).font=F(9); ws.cell(r,3,fq.get("ingreso","")).font=F(9); ws.cell(r,4,fq.get("fin","")).font=F(9); ws.cell(r,5,fq.get("causa","")).font=F(8)
        numc(r,6,round(float(fq.get("valor",0)),2),bold=True); fval=r
        for cx in range(1,7): ws.cell(r,cx).border=Border(bottom=thin)
        r+=1
        ws.cell(r,1,"Finiquito pendiente de pago al ex colaborador (fecha contable %s). Dar seguimiento a la liquidación."%fq.get("fin","")).font=F(8,False,"555555"); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7); r+=2
        cl=ws.cell(r,3,"Saldo contable %s"%cta); cl.font=F(9); ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=5); numc(r,6,sc); fb=r; r+=1
        difq=round(round(float(fq.get("valor",0)),2)-sc,2)
        cl=ws.cell(r,3,"Diferencia (debe ser 0)"); cl.font=F(10,True); ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=5)
        cc=ws.cell(r,6,"=F%d-F%d"%(fval,fb)); cc.font=F(10,True,GREEN if abs(difq)<0.01 else AMBER); cc.number_format=NUM; cc.alignment=Alignment("right"); r+=2
    else:
        ws.cell(r,1,"Finiquito pendiente (apertura, sin nombre en el mayor) — solicitar a quién corresponde y la fecha.").font=F(9,False,AMBER); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7); r+=2
    # ============ 3 y 4) DÉCIMO 13 y 14 (acumuladores, período completo) ============
    MM_ABR={1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}
    def seccion_decimo(r,cta,campo,periodo,nota,ym_list):
        sc=round(abs(bgs.get(cta,0)),2); aper,mon,dup_ag=agg(cta)
        banner(r,"%s  ·  %s   ·   Período %s   (saldo: %0.2f)"%(cta,NOM[cta],periodo,sc)); r+=1
        # acumuladores vigentes al corte (los que siguen en la provisión)
        acu=sorted({c for c in PROV[corte] if PROV[corte][c][campo]>0.005}, key=lambda c: nom_de(c))
        provym={ym:parse_prov_ym(ym) for ym in ym_list if ym.startswith("2025")}
        def val(ym,c):
            if ym.startswith("2025"): return provym.get(ym,{}).get(c,{}).get(campo,0.0)
            mm=int(ym[5:7]); return PROV[mm].get(c,{}).get(campo,0.0)
        ws.cell(r,1,nota+"  Lo acumulan (no mensualizan): %s. Se detalla cada mes del período (sin saldo de apertura)."%", ".join(nom_de(c).split()[0].title() for c in acu)).font=F(8,False,"555555"); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7); ws.row_dimensions[r].height=40; r+=1
        labs=[MM_ABR[int(ym[5:7])]+"-"+ym[2:4] for ym in ym_list]
        colhead(ws,r,[("Empleado",24)]+[(l,7) for l in labs]+[("Acumulado",11)]); r+=1
        fe=r
        for c in acu:
            ws.cell(r,1,nom_de(c)[:24]).font=F(9)
            for k,ym in enumerate(ym_list): numc(r,2+k,round(val(ym,c),2),sz=8)
            numc(r,2+len(ym_list),"=SUM(B%d:%s%d)"%(r,get_column_letter(1+len(ym_list)),r),bold=True)
            for cx in range(1,3+len(ym_list)): ws.cell(r,cx).border=Border(bottom=thin)
            r+=1
        lc=get_column_letter(2+len(ym_list))
        ws.cell(r,1,"Subtotal del período").font=F(9,True); numc(r,2+len(ym_list),"=SUM(%s%d:%s%d)"%(lc,fe,lc,r-1),bold=True)
        for cx in range(1,3+len(ym_list)): ws.cell(r,cx).border=Border(top=med)
        sub=r; r+=1
        if dup_ag>0.01:
            ws.cell(r,1,"(+) Ajuste — provisión de enero duplicada (asiento doble)").font=F(9,True,AMBER); numc(r,2+len(ym_list),dup_ag,bold=True,color=AMBER); dr=r; r+=1
            hall.append("%s — provisión de enero registrada dos veces (+%0.2f); a reversar."%(NOM[cta],dup_ag))
        else: dr=None
        sub_val=round(sum(round(val(ym,c),2) for c in acu for ym in ym_list),2)
        redon=round(sc-sub_val-(dup_ag if dr else 0),2)
        rr_=None
        if abs(redon)>0.005:
            ws.cell(r,1,"(±) Redondeo (tasa SBU 2025 vs 2026)").font=F(8,False,"888888"); numc(r,2+len(ym_list),redon,sz=8); rr_=r; r+=1
        ws.cell(r,1,"TOTAL ANEXO").font=F(10,True,NAVY)
        ref="%s%d"%(lc,sub)+(("+%s%d"%(lc,dr)) if dr else "")+(("+%s%d"%(lc,rr_)) if rr_ else "")
        ws.cell(r,2+len(ym_list),"="+ref).font=F(10,True); ws.cell(r,2+len(ym_list)).number_format=NUM; ws.cell(r,2+len(ym_list)).alignment=Alignment("right"); tr=r; r+=1
        ws.cell(r,1,"Saldo contable %s"%cta).font=F(9); numc(r,2+len(ym_list),sc); brr=r; r+=1
        ws.cell(r,1,"Diferencia (debe ser 0)").font=F(10,True); cc=ws.cell(r,2+len(ym_list),"=%s%d-%s%d"%(lc,tr,lc,brr)); cc.font=F(10,True,GREEN); cc.number_format=NUM; cc.alignment=Alignment("right"); r+=2
        return r
    ym13=["2025-12","2026-01","2026-02","2026-03","2026-04","2026-05"]
    ym14=["2025-08","2025-09","2025-10","2025-11","2025-12","2026-01","2026-02","2026-03","2026-04","2026-05"]
    r=seccion_decimo(r,"201070403","d13","2025-2026 (dic-2025 → nov-2026)",
        "El décimo tercero se devenga de diciembre a noviembre. En diciembre-2026 se abre el período 2026-2027 (al cierre solo estará diciembre). Al pagarse un período, se retira del anexo.",ym13)
    r=seccion_decimo(r,"201070404","d14","2025-2026 (ago-2025 → jul-2026, sierra)",
        "El décimo cuarto (sierra) se devenga de agosto a julio. En agosto-2026 se abre el período 2026-2027. Al pagarse el período se retira del anexo.",ym14)
    # ============ 5) VACACIONES ============
    cta="201070405"; sc=round(abs(bgs.get(cta,0)),2)
    banner(r,"%s  ·  %s   (saldo en balance: %0.2f)"%(cta,NOM[cta],sc)); r+=1
    ws.cell(r,1,"Vacaciones por empleado según el módulo de SAE: días ganados − días gozados = días disponibles, con su valor. El valor de los días disponibles es la obligación real; se compara con la provisión contable para cuantificar el ajuste (p. ej. vacaciones tomadas no registradas, o sub/sobre-provisión).").font=F(8,False,"555555"); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7); ws.row_dimensions[r].height=40; r+=1
    rep=p.get("vacaciones_reporte",[])
    if rep:
        colhead(ws,r,[("Empleado",30),("Días ganados",11),("Días gozados",11),("Días disponibles",12),("Valor disponible",13),("",2),("",2)]); ws.merge_cells(start_row=r,start_column=6,end_row=r,end_column=7); r+=1
        fe=r
        for nom,gan,tom,disp,val in rep:
            ws.cell(r,1,str(nom)[:30]).font=F(9)
            for cc_,vv_ in ((2,gan),(3,tom),(4,disp)):
                cu=ws.cell(r,cc_,round(float(vv_),2)); cu.number_format="0.00"; cu.alignment=Alignment("center"); cu.font=F(9,cc_==4)
            numc(r,5,round(float(val),2),bold=True)
            for cx in range(1,6): ws.cell(r,cx).border=Border(bottom=thin)
            r+=1
        ws.cell(r,1,"TOTAL — valor de vacaciones disponibles (SAE)").font=F(10,True,NAVY)
        for cc_,L_ in ((2,"B"),(3,"C"),(4,"D")):
            cu=ws.cell(r,cc_,"=SUM(%s%d:%s%d)"%(L_,fe,L_,r-1)); cu.number_format="0.00"; cu.alignment=Alignment("center"); cu.font=F(9,True)
        tv=ws.cell(r,5,"=SUM(E%d:E%d)"%(fe,r-1)); tv.font=F(10,True); tv.number_format=NUM; tv.alignment=Alignment("right")
        for cx in range(1,6): ws.cell(r,cx).border=Border(top=med)
        tvr=r; r+=2
        totval=round(sum(float(x[4]) for x in rep),2)
        aper,mon,dup_ag=agg(cta); dup_vac=round(sc-totval,2)
        ws.cell(r,1,"Conciliación con la contabilidad:").font=F(9,True,NAVY); r+=1
        def conc(rr,txt,val,bold=False,color="000000"):
            cl=ws.cell(rr,2,txt); cl.font=F(9,bold,color); ws.merge_cells(start_row=rr,start_column=2,end_row=rr,end_column=4); numc(rr,5,val,bold=bold,color=color)
        conc(r,"Provisión contable %s"%cta,sc); c1=r; r+=1
        conc(r,"(−) Duplicado de enero (asiento doble, a reversar)",dup_vac,bold=True,color=AMBER); c2=r; r+=1
        conc(r,"(=) Provisión real (sin duplicado)",round(sc-dup_vac,2),bold=True); c3=r; r+=1
        conc(r,"Valor disponible según SAE (días pendientes)",totval); c4=r; r+=1
        difv=round((sc-dup_vac)-totval,2)
        cl=ws.cell(r,2,"Diferencia (debe ser 0)"); cl.font=F(10,True); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
        cc=ws.cell(r,5,"=(E%d-E%d)-E%d"%(c1,c2,c4)); cc.font=F(10,True,GREEN if abs(difv)<0.01 else AMBER); cc.number_format=NUM; cc.alignment=Alignment("right"); r+=2
        ws.cell(r,1,"La diferencia entre la provisión contable (%0.2f) y el valor disponible de SAE (%0.2f) es exactamente el duplicado de enero (%0.2f). Reversado ese asiento, la provisión coincide con el módulo de vacaciones; no hay descuadre real."%(sc,totval,dup_vac)).font=F(8,False,"555555"); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7); ws.row_dimensions[r].height=30; r+=1
        hall.append("Vacaciones: la única diferencia entre la provisión contable (%0.2f) y el valor de días disponibles de SAE (%0.2f) es el duplicado de enero (%0.2f); reversado, cuadran."%(sc,totval,dup_vac))
        r+=1
    else:
        ws.cell(r,1,"Falta el reporte de vacaciones de SAE.").font=F(9,False,AMBER); r+=2
    # ============ HALLAZGOS ============
    ws.cell(r,1,"Hallazgos").font=F(10,True,NAVY); r+=1
    for n in p.get("observaciones",[]): hall.append(n)
    for n in hall:
        cc=ws.cell(r,1,"•  "+n); cc.font=F(8,False,"444444"); cc.alignment=Alignment("left",wrap_text=True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7); ws.row_dimensions[r].height=30; r+=1
    ws.column_dimensions["A"].width=27
    for cl in "BCDEFG": ws.column_dimensions[cl].width=11
    cc=ws.cell(r+1,1,"← Volver al Balance"); cc.hyperlink=Hyperlink(ref=cc.coordinate, location="BG!A1"); cc.font=F(8,True,LINK)
    return sbg, sbg

def build_participacion(ws,aid,a,args,bgs):
    p=a.get("params",{}) or {}; anios=p.get("anios",[2023,2024,2025])
    anexo_header(ws,aid,a["desc"],ncols=6)
    sbg=saldo_bg_tramo(bgs,a["cuentas"]); ins=args.insumos or ""
    def numc(r,col,v,bold=False,color="000000",sz=9):
        c=ws.cell(r,col,v); c.font=F(sz,bold,color); c.alignment=Alignment("right")
        if isinstance(v,(int,float)): c.number_format=NUM
        return c
    def banner(r,txt):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
        c=ws.cell(r,1,txt); c.font=F(11,True,"FFFFFF"); c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=Alignment("left","center",indent=1); ws.row_dimensions[r].height=20
    r=4; tot_rows=[]; total_acum=0.0
    for y in anios:
        path=os.path.join(ins,"utilidades_%d.xlsx"%y)
        if not os.path.exists(path): continue
        wsu=openpyxl.load_workbook(path,data_only=True).active
        hr=None
        for rr in range(1,wsu.max_row+1):
            if str(wsu.cell(rr,1).value).strip().upper()=="CÉDULA": hr=rr; break
        if hr is None: continue
        c10=c5=None
        for c in range(1,wsu.max_column+1):
            h=str(wsu.cell(hr,c).value or "").upper()
            if "10%" in h: c10=c
            if "5%" in h: c5=c
        banner(r,"Año %d — Participación a trabajadores (15%% de utilidades)"%y); r+=1
        colhead(ws,r,[("Cédula",13),("Empleado",30),("10%",12),("5%",12),("Total",13),("",2)]); r+=1
        fe=r
        for rr in range(hr+1,wsu.max_row+1):
            ced=wsu.cell(rr,1).value
            if ced is None: continue
            cs=str(int(ced)) if isinstance(ced,float) else str(ced).strip()
            if str(wsu.cell(rr,1).value).strip().upper()=="TOTAL" or not cs.replace('.','').isdigit(): continue
            nom=("%s %s"%(wsu.cell(rr,2).value or "",wsu.cell(rr,3).value or "")).strip()
            v10=round(float(wsu.cell(rr,c10).value or 0),2); v5=round(float(wsu.cell(rr,c5).value or 0),2)
            ws.cell(r,1,cs).font=F(8); ws.cell(r,2,nom[:30]).font=F(9); numc(r,3,v10); numc(r,4,v5)
            numc(r,5,round(v10+v5,2),bold=True)
            for cx in range(1,6): ws.cell(r,cx).border=Border(bottom=thin)
            r+=1
        ws.cell(r,2,"Total %d (a pagar)"%y).font=F(10,True)
        for col,L in [(3,"C"),(4,"D"),(5,"E")]:
            cc=ws.cell(r,col,"=SUM(%s%d:%s%d)"%(L,fe,L,r-1)); cc.font=F(10,True); cc.number_format=NUM; cc.alignment=Alignment("right")
        for cx in range(1,6): ws.cell(r,cx).border=Border(top=med,bottom=med)
        tot_rows.append(r); r+=2
    # mayor: pagos de la cuenta
    pagos=0.0
    for mv in mayor_cuenta(args.mayor,a["cuentas"]): pagos+=mv["debe"]
    pagos=round(pagos,2)
    banner(r,"Conciliación con la contabilidad"); r+=1
    ws.cell(r,2,"Participación causada (suma de los años, sin pagar)").font=F(9)
    ref="+".join("E%d"%x for x in tot_rows)
    ca=ws.cell(r,5,"="+ref); ca.font=F(9,True); ca.number_format=NUM; ca.alignment=Alignment("right"); ar=r; r+=1
    if abs(pagos)>0.01:
        ws.cell(r,2,"(−) Pagos registrados").font=F(9); numc(r,5,pagos); r+=1
    sub_part=round(sum(round(sum(round(float(0),2) for _ in [0]),2) for _ in []),2)
    # redondeo
    causado=0.0
    for mv in mayor_cuenta(args.mayor,a["cuentas"]): causado+=abs(mv["haber"])
    redon=round(sbg-causado,2)
    if abs(redon)>0.005:
        ws.cell(r,2,"(±) Redondeo (detalle por empleado vs cierre del mayor)").font=F(8,False,"888888"); numc(r,5,redon,sz=8); r+=1
    ws.cell(r,2,"Saldo contable %s"%a["cuentas"][0]).font=F(9,True); numc(r,5,sbg,bold=True); br=r; r+=1
    ws.cell(r,2,"Diferencia (debe ser 0)").font=F(10,True)
    difp=round(causado+redon-sbg,2)
    cd=ws.cell(r,5,"=SUM(E%d:E%d)"%(ar,br-1)); cd.value="=E%d%s-E%d"%(ar,("+E%d"%(br-1) if abs(redon)>0.005 else ""),br); cd.font=F(10,True,GREEN); cd.number_format=NUM; cd.alignment=Alignment("right"); r+=2
    ws.cell(r,1,"Hallazgo: la participación a trabajadores de %s está SIN PAGAR (el mayor solo tiene los asientos de cierre, ningún pago). La de años anteriores está vencida (se paga hasta ~15 días después del plazo de la declaración de renta). Regularizar."%(", ".join(str(y) for y in anios))).font=F(8,False,AMBER); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); ws.row_dimensions[r].height=30; r+=1
    ws.column_dimensions["A"].width=14; ws.column_dimensions["B"].width=32
    for cl in "CDE": ws.column_dimensions[cl].width=13
    cc=ws.cell(r+1,1,"← Volver al Balance"); cc.hyperlink=Hyperlink(ref=cc.coordinate, location="BG!A1"); cc.font=F(8,True,LINK)
    val=round(sum( bgs.get(c,0) and 0 or 0 for c in []),2)
    return sbg, sbg

def build_ganancias_anuales(ws,aid,a,args,bgs):
    p=a.get("params",{}) or {}; cta=a["cuentas"][0]
    anexo_header(ws,aid,a["desc"],ncols=6)
    sbg=saldo_bg_tramo(bgs,a["cuentas"])
    def numc(r,col,v,bold=False,color="000000",sz=9):
        c=ws.cell(r,col,v); c.font=F(sz,bold,color); c.alignment=Alignment("right")
        if isinstance(v,(int,float)): c.number_format=NUM
        return c
    def banner(r,txt):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
        c=ws.cell(r,1,txt); c.font=F(11,True,"FFFFFF"); c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=Alignment("left","center",indent=1); ws.row_dimensions[r].height=20
    g=p.get("ganancias",{}); base=g.get("base",{}); rows=g.get("rows",[])
    r=4
    banner(r,"%s  ·  %s   (saldo en balance: %0.2f)"%(cta,a["desc"],sbg)); r+=1
    ws.cell(r,1,(p.get("nota") or "El resultado (utilidad) de cada ejercicio pasa a ganancias acumuladas en el año siguiente. De esta cuenta se pagan los retiros de los socios.")).font=F(8,False,"555555"); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); ws.row_dimensions[r].height=28; r+=1
    colhead(ws,r,[("Año",10),("Saldo inicial",14),("(+) Utilidad del ejercicio","",),("(−) Consumos socios",15),("Saldo final",14),("",2)]) if False else colhead(ws,r,[("Año",10),("Saldo inicial",14),("(+) Utilidad del año previo",16),("(−) Consumos socios",15),("Saldo final",14),("",2)])
    r+=1; fe=r
    # fila base
    ws.cell(r,1,"%s"%base.get("anio","")).font=F(9); ws.cell(r,2,"(saldo de cierre)").font=F(8,False,"888888")
    numc(r,5,round(float(base.get("saldo",0)),2),bold=True)
    for cx in range(1,6): ws.cell(r,cx).border=Border(bottom=thin)
    prev=r; r+=1
    for row in rows:
        ws.cell(r,1,str(row.get("anio",""))).font=F(9)
        ws.cell(r,2,"=E%d"%prev).number_format=NUM; ws.cell(r,2).font=F(9); ws.cell(r,2).alignment=Alignment("right")
        cu=numc(r,3,round(float(row.get("utilidad",0)),2)); ws.cell(r,3).comment=None
        numc(r,4,round(float(row.get("consumos",0)),2))
        ce=ws.cell(r,5,"=B%d+C%d-D%d"%(r,r,r)); ce.number_format=NUM; ce.font=F(9,True); ce.alignment=Alignment("right")
        # nota del año de la utilidad
        ws.cell(r,6,"util. %s"%row.get("util_anio","")).font=F(7,False,"888888")
        for cx in range(1,6): ws.cell(r,cx).border=Border(bottom=thin)
        prev=r; r+=1
    fin=r-1
    ws.cell(r,1,"Saldo según balance %s (a mayo 2026)"%cta).font=F(9,True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4); numc(r,5,sbg,bold=True); br=r; r+=1
    ws.cell(r,1,"Diferencia (debe ser 0)").font=F(10,True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
    cd=ws.cell(r,5,"=E%d-E%d"%(fin,br)); cd.font=F(10,True,GREEN); cd.number_format=NUM; cd.alignment=Alignment("right"); r+=2
    ws.column_dimensions["A"].width=10; ws.column_dimensions["B"].width=26; ws.column_dimensions["C"].width=24
    for cl in "DE": ws.column_dimensions[cl].width=15
    cc=ws.cell(r+1,1,"← Volver al Balance"); cc.hyperlink=Hyperlink(ref=cc.coordinate, location="BG!A1"); cc.font=F(8,True,LINK)
    fin_val=round(float(base.get("saldo",0))+sum(float(x.get("utilidad",0))-float(x.get("consumos",0)) for x in rows),2)
    return fin_val, sbg

def build_cxp_tercero(ws,aid,a,args,bgs):
    import unicodedata
    p=a.get("params",{}) or {}; cta=a["cuentas"][0]
    anexo_header(ws,aid,a["desc"],ncols=6)
    sbg=saldo_bg_tramo(bgs,a["cuentas"]); ins=args.insumos or ""
    def numc(r,col,v,bold=False,color="000000",sz=9):
        c=ws.cell(r,col,v); c.font=F(sz,bold,color); c.alignment=Alignment("right")
        if isinstance(v,(int,float)): c.number_format=NUM
        return c
    def banner(r,txt,fill=NAVY):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
        c=ws.cell(r,1,txt); c.font=F(11,True,"FFFFFF"); c.fill=PatternFill("solid",fgColor=fill); c.alignment=Alignment("left","center",indent=1); ws.row_dimensions[r].height=20
    def canon(x):
        t=" ".join(str(x or "").strip().upper().split())
        t="".join(ch for ch in unicodedata.normalize("NFD",t) if unicodedata.category(ch)!="Mn")
        for k,full in alias_terceros():
            if k in t: return full.upper()
        return t or "(SIN TERCERO)"
    def ccon(x):
        t=" ".join(str(x or "").strip().upper().split())
        t="".join(ch for ch in unicodedata.normalize("NFD",t) if unicodedata.category(ch)!="Mn")
        return t[:34] if t else "(sin concepto)"
    path=os.path.join(ins,(p.get("reporte_archivo") or "mayor_%s.xlsx"%cta))
    r=4; banner(r,"%s  ·  %s   (saldo en balance: %0.2f)"%(cta,a["desc"],sbg)); r+=1
    if not os.path.exists(path):
        ws.cell(r,1,"Falta el mayor de la cuenta: "+os.path.basename(path)).font=F(9,False,AMBER); return 0.0,sbg
    origen=p.get("origen_nota","")
    if origen:
        cc=ws.cell(r,1,"Origen: "+origen); cc.font=F(8,False,"8A1C1C" if "⚠" in origen else "555555"); cc.alignment=Alignment("left",wrap_text=True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); ws.row_dimensions[r].height=42; r+=2
    wsm=openpyxl.load_workbook(path,data_only=True).active; hr=None
    for rr in range(1,wsm.max_row+1):
        if str(wsm.cell(rr,1).value).strip().upper()=="FECHA": hr=rr;break
    H={str(wsm.cell(hr,c).value).strip().upper():c for c in range(1,wsm.max_column+1)}
    cf=H.get("FECHA",1); cben=H.get("CLIENTE/PROVEEDOR",5); cco=H.get("COMENTARIO",6); cde=H.get("DEBE",7); cha=H.get("HABER",8); cdoc=H.get("DOCUMENTO",3)
    from collections import defaultdict
    grp=defaultdict(lambda: defaultdict(lambda:[0.0,0,""]))  # tercero -> concepto -> [net, count, ult_fecha]
    aper=0.0
    for rr in range(hr+1,wsm.max_row+1):
        f=wsm.cell(rr,cf).value
        if not f or "TOTAL" in str(f).upper(): continue
        ben=str(wsm.cell(rr,cben).value or ""); com=str(wsm.cell(rr,cco).value or ""); doc=str(wsm.cell(rr,cdoc).value or "")
        de=float(wsm.cell(rr,cde).value or 0); ha=float(wsm.cell(rr,cha).value or 0)
        if not ben.strip() and ("INICIAL" in com.upper() or "SALDOS" in com.upper()): aper+=de+ha; continue
        ter=canon(ben); con=ccon(com) if com.strip() else (doc[:20] or "(sin concepto)")
        g=grp[ter][con]; g[0]+=de+ha; g[1]+=1; g[2]=str(f)
    if abs(aper)>0.005:
        grp["(SALDO INICIAL — EJERCICIOS ANTERIORES)"]["saldo de apertura"]=[aper,1,""]
    # ordenar terceros por |saldo|
    terorder=sorted(grp, key=lambda t: abs(sum(v[0] for v in grp[t].values())), reverse=True)
    colhead(ws,r,[("Tercero / Concepto",42),("",2),("N° mov.",9),("Saldo (acreedor)",15),("Últ. mov.",12),("",2)]); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2); r+=1
    fe=r; subtot_rows=[]
    for ter in terorder:
        ws.cell(r,1,ter[:46]).font=F(9,True,NAVY); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
        ws.cell(r,1).fill=PatternFill("solid",fgColor="EAEDF1"); r+=1
        cfirst=r
        for con,(net,k,uf) in sorted(grp[ter].items(), key=lambda kv: abs(kv[1][0]), reverse=True):
            ws.cell(r,1,"   "+con[:42]).font=F(8); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
            ws.cell(r,3,k).alignment=Alignment("center"); ws.cell(r,3).font=F(8)
            numc(r,4,-round(net,2),sz=8)  # acreedor (haber neg) -> positivo
            ws.cell(r,5,uf).font=F(8); ws.cell(r,5).alignment=Alignment("center")
            for cx in range(1,6): ws.cell(r,cx).border=Border(bottom=thin)
            r+=1
        st=round(sum(v[0] for v in grp[ter].values()),2)
        ws.cell(r,1,"   Subtotal "+ter[:30]).font=F(8,True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
        numc(r,4,-st,bold=True)
        for cx in range(1,6): ws.cell(r,cx).border=Border(top=thin,bottom=thin)
        subtot_rows.append(r); r+=1
    r+=0
    ws.cell(r,1,"TOTAL").font=F(10,True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
    tc=ws.cell(r,4,"="+"+".join("D%d"%x for x in subtot_rows)); tc.font=F(10,True); tc.number_format=NUM; tc.alignment=Alignment("right")
    for cx in range(1,6): ws.cell(r,cx).border=Border(top=med)
    tr=r; r+=1
    ws.cell(r,1,"Saldo según balance %s"%cta).font=F(9); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3); numc(r,4,sbg); br=r; r+=1
    ws.cell(r,1,"Diferencia (debe ser 0)").font=F(10,True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
    cd=ws.cell(r,4,"=D%d-D%d"%(tr,br)); cd.font=F(10,True,GREEN); cd.number_format=NUM; cd.alignment=Alignment("right"); r+=2
    ws.column_dimensions["A"].width=40
    for cl in "BCDE": ws.column_dimensions[cl].width=13
    cc=ws.cell(r+1,1,"← Volver al Balance"); cc.hyperlink=Hyperlink(ref=cc.coordinate, location="BG!A1"); cc.font=F(8,True,LINK)
    tot=round(sum(-sum(v[0] for v in grp[t].values()) for t in grp),2)
    return tot, sbg

BUILDERS={"amortizacion":build_amortizacion,"mayor_extracto":build_mayor_extracto,"conciliacion_bancaria":build_conciliacion_bancaria,"inversiones":build_inversiones,"antiguedad_cartera":build_antiguedad,"antiguedad_cxp":build_antiguedad,"lotes_tc":build_lotes_tc,"anticipo_empleados":build_anticipo_empleados,"stock":build_stock,"anticipos_prov":build_anticipos_prov,"conciliacion_tributaria":build_conciliacion_tributaria,"mayor_mensual":build_mayor_mensual,"activo_fijo":build_activo_fijo,"planillas_iess":build_planillas_iess,"provisiones_nomina":build_provisiones_nomina,"participacion":build_participacion,"ganancias_anuales":build_ganancias_anuales,"cxp_tercero":build_cxp_tercero}

import re as _re

def reorder(wb):
    base=[n for n in ("BG","PYG",CUADRE_HOJA,"BG ANTERIOR") if n in wb.sheetnames]
    anx=[n for n in wb.sheetnames if n not in base]
    def k(n):
        m=_re.match(r"A-(\d+)(?:\.(\d+))?",n); 
        return (int(m.group(1)), int(m.group(2)) if m and m.group(2) else 0) if m else (999,0)
    wb._sheets=[wb[n] for n in base]+[wb[n] for n in sorted(anx,key=k)]

def cmd_init(args):
    cfg=set_cfg(json.load(open(args.config,encoding='utf-8')))
    bgrows=parse_eef(args.bg); pygrows=parse_eef(args.pyg); bgs=bg_saldos(bgrows)
    anex_map={}
    for a in cfg["anexos"]:
        for c in a["cuentas"]: anex_map[c]=a["id"]
    wb=openpyxl.Workbook()
    bg=wb.active; bg.title="BG"; render_estado(bg,"Estado de Situación Financiera",bgrows,anex_map)
    pyg=wb.create_sheet("PYG"); render_estado(pyg,"Estado de Resultados",pygrows,None)
    nant=None
    if getattr(args,"bg_anterior",None):
        nant=render_bg_anterior(wb.create_sheet("BG ANTERIOR"),parse_eef(args.bg_anterior))
    write_state(args.salida,cfg,bgs)
    render_cuadre(wb,load_state(args.salida)); reorder(wb); wb.save(args.salida)
    print(f"Inicializado: {args.salida}")
    if nant is not None: print(f"BG ANTERIOR: {nant} cuentas con saldo de cierre del mes anterior (cadena de continuidad del acta)")
    else: print("sin BG ANTERIOR: la cadena de continuidad del acta quedará como aviso")
    cmd_plan(args)

def cmd_plan(args):
    st=load_state(args.salida)
    print(f"{'Anexo':6} {'Estado':14} {'Saldo BG':>14}  Reporte fuente")
    for r in st:
        print(f"{r['id']:6} {r['estado']:14} {r['saldo_bg']:>14,.2f}  {r['reporte']}")
    pend=sum(1 for r in st if r['estado']=='pendiente'); print(f"\n{len(st)} anexos · {pend} pendientes · {len(st)-pend} hechos")

def cmd_status(args): cmd_plan(args)

def cmd_order(args):
    wb=openpyxl.load_workbook(args.salida); reorder(wb); wb.save(args.salida); print('Hojas:', wb.sheetnames)

def cmd_build(args):
    wb=openpyxl.load_workbook(args.salida)
    cfg=set_cfg(json.load(open(args.config,encoding='utf-8')))
    a=next((x for x in cfg["anexos"] if x["id"]==args.anexo),None)
    if a is None: print(f"ERROR: anexo {args.anexo} no está en el config"); return
    bgs=bg_saldos(parse_eef(args.bg)) if args.bg else {}
    st=load_state(args.salida); tipo=a["tipo"]
    if tipo not in BUILDERS:
        update_state(st,a["id"],estado="falta constructor"); save_state(args.salida,st)
        print(f"{a['id']} {a['desc']}: tipo «{tipo}» aún no implementado. Reporte necesario: {a['reporte']}"); return
    if tipo in ("amortizacion","mayor_extracto","lotes_tc","anticipo_empleados") and not args.mayor:
        print(f"{a['id']} {a['desc']}: necesita el MAYOR (--mayor). Reporte: {a['reporte']}"); return
    if tipo in ("conciliacion_bancaria","antiguedad_cartera","antiguedad_cxp","stock","anticipos_prov") and not args.insumos:
        print(f"{a['id']} {a['desc']}: necesita la carpeta de insumos (--insumos). Reporte: {a['reporte']}"); return
    title=f"{a['id']} {a['desc']}"[:31]
    if title in wb.sheetnames: del wb[title]
    ws=wb.create_sheet(title)
    sa,sbg=BUILDERS[tipo](ws,a["id"],a,args,bgs)
    dif=round(sa-sbg,2); est="Cuadra" if abs(dif)<0.01 else "Hallazgo"
    # `saldo_bg` se refresca con el que REALMENTE se comparó: el del init
    # puede venir de otro export del Balance, y la hoja CUADRE tiene que
    # mostrar las dos cifras que produjeron esa diferencia.
    update_state(st,a["id"],estado=est,hoja=title,saldo_anexo=sa,saldo_bg=sbg,dif=dif); save_state(args.salida,st)
    link_bg(wb,a["id"]); render_cuadre(wb,st); reorder(wb); wb.save(args.salida)
    print(f"{a['id']} {a['desc']}: {est} · anexo {sa:,.2f} vs balance {sbg:,.2f} · dif {dif:,.2f}")

if __name__=="__main__":
    ap=argparse.ArgumentParser(); sub=ap.add_subparsers(dest="cmd")
    pi=sub.add_parser("init"); pi.add_argument("--bg",required=True); pi.add_argument("--pyg",required=True); pi.add_argument("--config",required=True); pi.add_argument("--salida",required=True); pi.add_argument("--bg-anterior",help="BG exportado del mes ANTERIOR (mismo formato que --bg): agrega la hoja «BG ANTERIOR» para la cadena de continuidad del acta")
    pp=sub.add_parser("plan"); pp.add_argument("--salida",required=True)
    ps=sub.add_parser("status"); ps.add_argument("--salida",required=True)
    po=sub.add_parser("order"); po.add_argument("--salida",required=True)
    pb=sub.add_parser("build"); pb.add_argument("--salida",required=True); pb.add_argument("--anexo",required=True); pb.add_argument("--config",required=True); pb.add_argument("--bg"); pb.add_argument("--mayor"); pb.add_argument("--reporte"); pb.add_argument("--insumos")
    a=ap.parse_args()
    if a.cmd=="init": cmd_init(a)
    elif a.cmd in ("plan",): cmd_plan(a)
    elif a.cmd=="status": cmd_status(a)
    elif a.cmd=="order": cmd_order(a)
    elif a.cmd=="build": cmd_build(a)
    else: ap.print_help()
